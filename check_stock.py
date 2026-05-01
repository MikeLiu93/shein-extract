"""
Lightweight stock checker: reads the input Excel, visits each Done URL,
extracts stock info only (no image download), and writes back to the
Stock and Last Checked columns.

Usage:
    python check_stock.py                          # scan submitted/ for .xlsx
    python check_stock.py "path/to/file.xlsx"      # specific file

Only processes rows where Status == "Done".
Writes:
    Column 5 (Stock): "In Stock (123)" / "Low Stock (5)" / "Sold Out" / "Delisted"
    Column 6 (Last Checked): YYYY-MM-DD
"""

import argparse
import json
import logging
import os
import re
import sys
import time
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook

from config import SUBMITTED_DIR, BACKUP_DIR, INPUT_FILENAME

logger = logging.getLogger("check_stock")
DEFAULT_XLSX = INPUT_FILENAME
DEBUG_LOG_DIR = Path(__file__).resolve().parent / "debug_logs"

LOW_STOCK_THRESHOLD = 15
DELAY_BETWEEN_PAGES = 4   # slightly longer than scraper to reduce risk
BATCH_SIZE = 200           # pause between batches
BATCH_PAUSE_SEC = 60       # 1 min pause between batches
PAGE_WAIT_SEC = 15         # max wait for goods_sn

# Minimal JS: extract only stock from sku_prices
_JS_STOCK = r"""
(function() {
    try {
        var gb = window.gbRawData;
        if (!gb || !gb.modules || !gb.modules.saleAttr) return null;
        var multi = gb.modules.saleAttr.multiLevelSaleAttribute;
        if (!multi || !multi.sku_list) return null;
        var total = 0;
        var skus = multi.sku_list;
        for (var i = 0; i < skus.length; i++) {
            total += (skus[i].stock || 0);
        }
        return total;
    } catch(e) { return null; }
})()
"""


def setup_logging():
    DEBUG_LOG_DIR.mkdir(parents=True, exist_ok=True)
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    log_path = DEBUG_LOG_DIR / f"stock_{ts}.log"
    fmt = logging.Formatter("%(asctime)s [%(levelname)s] %(message)s")

    root = logging.getLogger("check_stock")
    root.handlers.clear()
    root.setLevel(logging.DEBUG)

    fh = logging.FileHandler(log_path, encoding="utf-8")
    fh.setLevel(logging.DEBUG)
    fh.setFormatter(fmt)
    root.addHandler(fh)

    sh = logging.StreamHandler(sys.stdout)
    sh.setLevel(logging.INFO)
    sh.setFormatter(fmt)
    root.addHandler(sh)

    logger.info("Log: %s", log_path)


def backup_excel(xlsx_path: Path) -> None:
    """Backup Excel to personal Drive before reading."""
    import shutil
    BACKUP_DIR.mkdir(parents=True, exist_ok=True)
    today = datetime.now().strftime("%Y%m%d")
    dest = BACKUP_DIR / f"{xlsx_path.stem}_{today}.xlsx"
    try:
        shutil.copy2(xlsx_path, dest)
        logger.info("Backup: %s → %s", xlsx_path.name, dest)
    except Exception as e:
        logger.warning("Backup failed (continuing): %s", e)


def safe_save(wb, xlsx_path: Path) -> None:
    """Save workbook. If locked, save as copy with '2' suffix."""
    try:
        wb.save(xlsx_path)
    except PermissionError:
        alt = xlsx_path.with_stem(xlsx_path.stem + "2")
        logger.warning("Cannot save to %s (locked), saving to %s", xlsx_path.name, alt.name)
        wb.save(alt)


def _stock_label(total_stock: int | None) -> str:
    if total_stock is None:
        return "Unknown"
    if total_stock == 0:
        return "Sold Out"
    if total_stock <= LOW_STOCK_THRESHOLD:
        return f"Low Stock ({total_stock})"
    return f"In Stock ({total_stock})"


def check_stock_excel(xlsx_path: Path) -> None:
    # v3.4.1+: navigate via PUT /json/new?<url> (fresh tab per URL, no Referer).
    # Old single-tab + window.location.href approach was removed.
    from shein_scraper import (
        _navigate_and_wait, _close_tab, _run_js, _ws_url_for_id,
        CDP_PORT, _ensure_chrome,
    )

    logger.info("Opening: %s", xlsx_path.name)
    wb = load_workbook(xlsx_path)

    # Ensure Chrome is running once for the whole file (not per-sheet)
    _ensure_chrome(CDP_PORT)
    today = datetime.now().strftime("%Y-%m-%d")

    for ws_name in wb.sheetnames:
        ws = wb[ws_name]
        store = ws_name.strip()
        logger.info("Sheet: %s", store)

        # Ensure headers in columns 5 and 6
        if ws.cell(1, 5).value != "Stock":
            ws.cell(1, 5).value = "Stock"
        if ws.cell(1, 6).value != "Last Checked":
            ws.cell(1, 6).value = "Last Checked"

        # Collect rows: Status == "Done", needs stock check
        pending = []
        for row in range(2, ws.max_row + 1):
            status = str(ws.cell(row, 4).value or "").strip()
            url = ws.cell(row, 2).value
            seq = ws.cell(row, 1).value
            if status == "Done" and url:
                pending.append((row, seq, str(url).strip()))

        if not pending:
            logger.info("  No Done rows to check in '%s'", store)
            continue

        logger.info("  %d URL(s) to check stock", len(pending))

        for idx, (row, seq, url) in enumerate(pending):
            logger.info("  [%d/%d] seq %s: %s", idx + 1, len(pending),
                        seq, url[:60])

            stock_val = None
            tab_id = None
            try:
                ws_url, tab_id = _navigate_and_wait(CDP_PORT, url)

                # Check for Oops / delisted / NO_DATA placeholder
                deadline = time.monotonic() + PAGE_WAIT_SEC
                while time.monotonic() < deadline:
                    try:
                        ws_url = _ws_url_for_id(CDP_PORT, tab_id)
                        page_check = _run_js(ws_url, """
                            (function() {
                                if (document.body && (
                                    document.body.innerText.includes('Oops') ||
                                    document.querySelector('.page-not-found, [class*="not-found"]')
                                )) return 'OOPS';
                                if (document.title && document.title.includes('[goods_name]'))
                                    return 'NO_DATA';
                                return 'OK';
                            })()
                        """)
                        if page_check == "OOPS":
                            stock_val = -1  # delisted
                            break
                        if page_check == "NO_DATA":
                            time.sleep(1)
                            continue

                        s = _run_js(ws_url, _JS_STOCK)
                        if s is not None:
                            stock_val = int(s)
                            break
                    except Exception:
                        pass
                    time.sleep(1)
            except Exception as e:
                logger.warning("    nav error: %s", e)
            finally:
                if tab_id is not None:
                    _close_tab(CDP_PORT, tab_id)

            # Write result
            if stock_val == -1:
                label = "Delisted"
            elif stock_val is not None:
                label = _stock_label(stock_val)
            else:
                label = "Unknown"

            ws.cell(row, 5).value = label
            ws.cell(row, 6).value = today
            logger.info("    → %s", label)

            # Save every 10 rows
            if (idx + 1) % 10 == 0:
                safe_save(wb, xlsx_path)
                logger.info("  Saved progress (%d/%d)", idx + 1, len(pending))

            # Batch pause
            if (idx + 1) % BATCH_SIZE == 0 and idx + 1 < len(pending):
                logger.info("  Batch pause %ds...", BATCH_PAUSE_SEC)
                safe_save(wb, xlsx_path)
                time.sleep(BATCH_PAUSE_SEC)

            time.sleep(DELAY_BETWEEN_PAGES)

        safe_save(wb, xlsx_path)
        logger.info("  Done: %s (%d checked)", store, len(pending))

    wb.close()
    logger.info("All done.")


def main():
    parser = argparse.ArgumentParser(description="Shein stock checker")
    parser.add_argument("file", nargs="?", default=None,
                        help="Path to .xlsx (default: Shein Submited Links.xlsx)")
    args = parser.parse_args()

    setup_logging()

    if args.file:
        files = [Path(args.file)]
    else:
        default = SUBMITTED_DIR / DEFAULT_XLSX
        if default.exists():
            files = [default]
        else:
            logger.error("Default file not found: %s", default)
            return

    for f in files:
        backup_excel(f)
        try:
            check_stock_excel(f)
        except Exception as e:
            logger.exception("Error: %s", e)

    logger.info("Stock check complete.")


if __name__ == "__main__":
    main()
