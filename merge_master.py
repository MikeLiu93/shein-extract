"""
Per-store master merger
=======================

把 `Listing - completed 2nd/{store}/` 下所有 daily Excel
（`{store}-{seq_min}-{seq_max}-{date}.xlsx`）合并成一个总表
`{store}_master.xlsx`，保存在同一店铺文件夹下。

合并规则：
  - 按 seq 去重，同一个 seq 出现多次时：
    * 优先选成功的（SKU 非空 + price > 0）
    * 同等成功度时选最新日期（从文件名末尾的 YYYYMMDD 提取）
  - 主行 + 它后面的变体子行（A 列为空）作为一组保留/丢弃
  - 失败行（红色背景）也会进总表，但价格/SKU 为空，便于追溯
  - 图片从 {store}/{seq}/img_001.webp 重新插入（不依赖原 Excel 内嵌图片）

用法：
  python merge_master.py C4              合并店铺 C4
  python merge_master.py C4 --dry-run    只预览不写入
  python merge_master.py                 交互式输入店铺代号

不会动任何 daily Excel，只生成新的 {store}_master.xlsx。
"""

import argparse
import logging
import re
import sys
from copy import copy
from datetime import datetime
from pathlib import Path

# Force UTF-8 stdout/stdin/stderr on Windows when launched from a .cmd
# whose codepage was just switched to 65001 (chcp 65001). Python sometimes
# guesses cp936 from the registry — this ensures Chinese prints/inputs work.
try:
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")
    sys.stderr.reconfigure(encoding="utf-8", errors="replace")
    sys.stdin.reconfigure(encoding="utf-8", errors="replace")
except (AttributeError, ValueError):
    pass

from openpyxl import Workbook, load_workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.utils import get_column_letter

from config import OUTPUT_ROOT_2ND

logger = logging.getLogger("merge_master")

PICTURE_COL = 4
PICTURE_MAX_HEIGHT_PX = 168


def setup_logging() -> None:
    fmt = logging.Formatter("%(asctime)s [%(levelname)s] %(message)s")
    logger.handlers.clear()
    logger.setLevel(logging.INFO)
    sh = logging.StreamHandler(sys.stdout)
    sh.setLevel(logging.INFO)
    sh.setFormatter(fmt)
    logger.addHandler(sh)


def _parse_filename_date(stem: str) -> "datetime | None":
    """从文件名末尾提取 YYYYMMDD：'C4-1-50-20260415' → datetime(2026,4,15)"""
    m = re.search(r"(\d{8})$", stem)
    if not m:
        return None
    try:
        return datetime.strptime(m.group(1), "%Y%m%d")
    except ValueError:
        return None


def _is_master_filename(stem: str) -> bool:
    s = stem.lower()
    return s.endswith("_master") or s.endswith("_master2")


def find_daily_excels(store_dir: Path, store_code: str) -> list[Path]:
    """找出所有 daily excel，按文件名日期排序（旧到新）。"""
    daily: list[Path] = []
    for f in store_dir.glob(f"{store_code}-*.xlsx"):
        if _is_master_filename(f.stem):
            continue
        if "$" in f.name or f.name.startswith("~"):  # Excel temp/lock files
            continue
        daily.append(f)
    daily.sort(key=lambda p: (_parse_filename_date(p.stem) or datetime.min, p.name))
    return daily


def _add_picture(ws, row: int, col: int, image_path: Path) -> float:
    """
    把图片插入指定单元格，返回行高（pt）。逻辑与 shein_scraper._add_picture_to_cell
    保持一致，所以总表里图片大小和现有 daily Excel 视觉一致。
    """
    if not image_path or not image_path.is_file():
        return 0
    try:
        xl_img = XLImage(str(image_path))
    except Exception as e:
        logger.debug("  cannot load image %s: %s", image_path, e)
        return 0
    ow, oh = max(1, int(xl_img.width)), max(1, int(xl_img.height))
    letter = get_column_letter(col)
    dim = ws.column_dimensions.get(letter)
    wchars = float(dim.width) if dim and dim.width else 14
    max_w = max(24, int((wchars * 7.0 + 5.0) * 0.90))
    max_h = PICTURE_MAX_HEIGHT_PX
    scale = min(max_w / ow, max_h / oh, 1.0)
    xl_img.width = max(1, int(ow * scale))
    xl_img.height = max(1, int(oh * scale))
    ws.add_image(xl_img, f"{letter}{row}")
    return min(200.0, max(40.0, xl_img.height * (72.0 / 96.0) + 10.0))


def _read_groups(ws, xlsx_path: Path) -> list[dict]:
    """
    扫描 worksheet 的所有数据行，把"主行 + 后续变体子行"打包成 group。

    返回：[
      {
        "seq": int,
        "source_file": Path,
        "source_date": datetime|None,
        "ws": worksheet,        # 持有引用，调用者必须让 wb 别被 GC
        "main_row": int,
        "rows": list[int],      # 主行 + 子行的源行号
        "is_success": bool,     # SKU 非空 + price > 0 → True
      }, ...
    ]
    """
    file_date = _parse_filename_date(xlsx_path.stem)
    groups: list[dict] = []
    current: dict | None = None

    if ws.max_row < 2:
        return groups

    for r in range(2, ws.max_row + 1):
        no_val = ws.cell(r, 1).value
        seq_num: int | None = None
        if no_val is not None and str(no_val).strip() != "":
            try:
                seq_num = int(no_val)
            except (ValueError, TypeError):
                seq_num = None

        if seq_num is not None:
            # 主行：开新 group
            sku_val = ws.cell(r, 3).value
            sku = str(sku_val).strip() if sku_val else ""
            try:
                price_val = ws.cell(r, 5).value
                price = float(price_val) if price_val not in (None, "") else 0.0
            except (ValueError, TypeError):
                price = 0.0
            is_success = bool(sku) and price > 0

            current = {
                "seq": seq_num,
                "source_file": xlsx_path,
                "source_date": file_date,
                "ws": ws,
                "main_row": r,
                "rows": [r],
                "is_success": is_success,
            }
            groups.append(current)
        elif current is not None:
            # 变体子行：归到上一个 group
            current["rows"].append(r)
        # else: 没主行就出现的孤儿子行 —— 跳过（理论上不会发生）

    return groups


def _pick_winner(groups_for_seq: list[dict]) -> dict:
    """
    同一个 seq 多个 group 时选最优：
      1. 成功优先
      2. 同等成功度 → 最新日期优先
      3. 还相同 → 第一个出现的（稳定排序）
    """
    def key(g: dict):
        ts = g["source_date"].timestamp() if g["source_date"] else 0
        return (-int(g["is_success"]), -ts)

    return sorted(groups_for_seq, key=key)[0]


def merge_store(store_code: str, store_dir: Path, dry_run: bool = False) -> bool:
    """
    合并一个店铺的所有 daily excel。返回 True 表示成功（即使 dry-run 也算成功）。
    """
    daily = find_daily_excels(store_dir, store_code)
    if not daily:
        logger.warning("  No daily excels found for '%s' under %s", store_code, store_dir)
        return False

    logger.info("Found %d daily excel(s):", len(daily))
    for f in daily:
        logger.info("  - %s", f.name)

    # 先用最后一个（最新）daily 当 header 模板
    template_wb = load_workbook(str(daily[-1]))
    template_ws = template_wb.active
    col_count = template_ws.max_column

    # 读所有 daily 的 group。注意：必须保持 wb 引用直到拷贝完，否则 ws.cell 会失效
    open_wbs: list = []
    all_groups: list[dict] = []
    for xlsx_path in daily:
        try:
            wb = load_workbook(str(xlsx_path))
        except Exception as e:
            logger.warning("  cannot open %s: %s — skip", xlsx_path.name, e)
            continue
        ws = wb.active
        col_count = max(col_count, ws.max_column)
        open_wbs.append(wb)
        groups = _read_groups(ws, xlsx_path)
        all_groups.extend(groups)
        logger.debug("  read %d group(s) from %s", len(groups), xlsx_path.name)

    if not all_groups:
        logger.warning("  no data rows found across %d daily file(s)", len(daily))
        return False

    # 按 seq 分组
    by_seq: dict[int, list[dict]] = {}
    for g in all_groups:
        by_seq.setdefault(g["seq"], []).append(g)

    # 选 winner
    winners: dict[int, dict] = {}
    for seq in sorted(by_seq.keys()):
        candidates = by_seq[seq]
        winner = _pick_winner(candidates)
        winners[seq] = winner
        if len(candidates) > 1:
            srcs = ", ".join(
                f"{c['source_file'].name}({'OK' if c['is_success'] else 'FAIL'})"
                for c in candidates
            )
            tag = "success" if winner["is_success"] else "best available"
            logger.info(
                "  seq %d: %d candidates [%s] → keep from %s (%s)",
                seq, len(candidates), srcs, winner["source_file"].name, tag,
            )

    n_total = len(winners)
    n_success = sum(1 for g in winners.values() if g["is_success"])
    n_failed = n_total - n_success
    logger.info(
        "Total: %d unique seq(s) — success=%d, failed/delisted=%d",
        n_total, n_success, n_failed,
    )

    if dry_run:
        logger.info("[DRY RUN] No file written. Run without --dry-run to actually write.")
        return True

    # 写新 master Excel
    new_wb = Workbook()
    new_ws = new_wb.active
    new_ws.title = "Shein Products"

    # 拷贝 header 格式
    for c in range(1, col_count + 1):
        src = template_ws.cell(1, c)
        dst = new_ws.cell(1, c, src.value)
        dst.font = copy(src.font)
        dst.fill = copy(src.fill)
        dst.alignment = copy(src.alignment)
        dst.border = copy(src.border)
        letter = get_column_letter(c)
        src_dim = template_ws.column_dimensions.get(letter)
        if src_dim and src_dim.width:
            new_ws.column_dimensions[letter].width = src_dim.width

    new_ws.row_dimensions[1].height = template_ws.row_dimensions[1].height or 24
    new_ws.freeze_panes = "A2"
    new_ws.auto_filter.ref = f"A1:{get_column_letter(col_count)}1"

    # 按 seq 升序写数据行
    dest_row = 2
    img_count = 0
    for seq in sorted(winners.keys()):
        winner = winners[seq]
        src_ws = winner["ws"]

        for src_r in winner["rows"]:
            # 拷贝所有列的值 + 格式
            for c in range(1, col_count + 1):
                src = src_ws.cell(src_r, c)
                dst = new_ws.cell(dest_row, c, src.value)
                dst.font = copy(src.font)
                if src.fill and src.fill.patternType:
                    dst.fill = copy(src.fill)
                dst.alignment = copy(src.alignment)
                dst.border = copy(src.border)
                if src.number_format:
                    dst.number_format = src.number_format

            # 行高
            src_h = src_ws.row_dimensions[src_r].height
            if src_h:
                new_ws.row_dimensions[dest_row].height = src_h

            # 图片：只在主行 + 成功 seq 上插，从磁盘 seq 文件夹读
            if src_r == winner["main_row"] and winner["is_success"]:
                img_path = store_dir / str(seq) / "img_001.webp"
                if not img_path.is_file():
                    # fallback：找 seq 文件夹里任意 img_*
                    seq_dir = store_dir / str(seq)
                    if seq_dir.is_dir():
                        any_img = sorted(seq_dir.glob("img_*.*"))
                        if any_img:
                            img_path = any_img[0]
                if img_path.is_file():
                    pt = _add_picture(new_ws, dest_row, PICTURE_COL, img_path)
                    if pt > 0:
                        img_count += 1
                        cur_h = new_ws.row_dimensions[dest_row].height or 18
                        new_ws.row_dimensions[dest_row].height = max(cur_h, pt)

            dest_row += 1

    # 保存（被锁就改名 ...master2.xlsx）
    out_path = store_dir / f"{store_code}_master.xlsx"
    try:
        new_wb.save(str(out_path))
    except PermissionError:
        alt = store_dir / f"{store_code}_master2.xlsx"
        logger.warning("  cannot save to %s (locked), saving to %s", out_path.name, alt.name)
        new_wb.save(str(alt))
        out_path = alt

    logger.info(
        "Saved: %s (%d data rows, %d images)",
        out_path.name, dest_row - 2, img_count,
    )
    return True


def main():
    parser = argparse.ArgumentParser(
        description="Merge a store's daily excels into one master xlsx",
    )
    parser.add_argument("store", nargs="?",
                        help="Store code (e.g. C4). Prompts if omitted.")
    parser.add_argument("--dry-run", action="store_true",
                        help="Preview merge without writing the master file.")
    args = parser.parse_args()

    setup_logging()

    print("============================================")
    print("  Shein 总表合并工具 (merge_master)")
    print("============================================")
    print()

    store_code = (args.store or "").strip()
    if not store_code:
        try:
            store_code = input("请输入店铺代号（例如 C4）: ").strip()
        except EOFError:
            store_code = ""
    if not store_code:
        print()
        logger.error("[错误] 未输入店铺代号，退出。")
