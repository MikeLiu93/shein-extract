"""
Merge Store Reports
===================
按员工 + 店铺代号，合并 Listing - completed 下的 shein_products_*.xlsx。
去重依据：第一列 No.（序列号），保留首次出现的行。
图片从磁盘上的 seq 文件夹 (img_001.webp) 直接插入，确保位置正确。

用法：
    python merge_store_reports.py          # 合并所有员工所有店铺
    python merge_store_reports.py LUMEI    # 只合并 LUMEI
    python merge_store_reports.py LUMEI C2 # 只合并 LUMEI 的 C2 店铺
"""

import argparse
import re
from collections import defaultdict
from copy import copy
from datetime import date, datetime, timedelta
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.utils import get_column_letter

COMPLETED_ROOT = Path(r"D:\共享云端硬盘\02 希音\Auto Pipeline\Listing - completed")
KNOWN_EMPLOYEES = ["NA", "TT", "YAN", "ZQW", "LUMEI"]
PICTURE_COL = 4
PICTURE_MAX_HEIGHT_PX = 168


def _get_week_range(ref_date: date = None) -> tuple[date, date]:
    """返回 ref_date 所在周的周一和周日。"""
    d = ref_date or date.today()
    monday = d - timedelta(days=d.weekday())
    sunday = monday + timedelta(days=6)
    return monday, sunday


def _parse_folder_date(folder_name: str) -> "date | None":
    """从文件夹名开头提取日期：'20260410 - NA - L8 - 79-88' → date(2026,4,10)。"""
    m = re.match(r'^(\d{8})', folder_name)
    if m:
        try:
            return datetime.strptime(m.group(1), "%Y%m%d").date()
        except ValueError:
            pass
    return None


def _parse_store_code(folder_name: str, employee: str) -> str:
    name = folder_name
    name = re.sub(r'^\d{8}\s*[-–]?\s*', '', name).strip()
    name = re.sub(rf'^{re.escape(employee)}\s*[-–]\s*', '', name, flags=re.I).strip()
    name = re.sub(r'^\s*[-–]\s*', '', name).strip()
    parts = re.split(r'\s*[-–]\s*', name)
    store = parts[0].strip() if parts else ""
    if store and re.match(r'^\d+$', store):
        return ""
    m = re.match(r'^([A-Za-z]+\d*)', store)
    return m.group(1).upper() if m else ""


def _find_xlsx_in_folder(folder: Path) -> list[Path]:
    candidates = [f for f in folder.glob("shein_products_*.xlsx")
                  if "merged" not in f.stem.lower() and "2nd run" not in f.stem.lower()]
    candidates += [f for f in folder.glob("*2nd run*.xlsx")]
    return candidates if candidates else []


def _find_first_image(folder: Path, seq_num) -> "Path | None":
    """在 seq 文件夹（如 79/img_001.webp）中找到第一张图片。"""
    if seq_num is None or str(seq_num).strip() == "":
        return None
    seq_dir = folder / str(int(seq_num))
    if not seq_dir.is_dir():
        # 也试试 -2 后缀
        seq_dir2 = folder / f"{int(seq_num)}-2"
        if seq_dir2.is_dir():
            seq_dir = seq_dir2
        else:
            return None
    img = seq_dir / "img_001.webp"
    if img.is_file():
        return img
    # fallback: 找任何 img_* 文件
    imgs = sorted(seq_dir.glob("img_*.*"))
    return imgs[0] if imgs else None


def _add_picture(ws, row: int, col: int, image_path: Path) -> float:
    """插入图片到指定单元格，返回行高 pt。与 shein_scraper 一致。"""
    if not image_path or not image_path.is_file():
        return 0
    try:
        xl_img = XLImage(str(image_path))
    except Exception:
        return 0
    ow, oh = max(1, int(xl_img.width)), max(1, int(xl_img.height))
    # 计算列内宽度
    letter = get_column_letter(col)
    dim = ws.column_dimensions.get(letter)
    wchars = float(dim.width) if dim and dim.width else 14
    max_w = max(24, int((wchars * 7.0 + 5.0) * 0.90))
    max_h = PICTURE_MAX_HEIGHT_PX
    scale = min(max_w / ow, max_h / oh, 1.0)
    xl_img.width = max(1, int(ow * scale))
    xl_img.height = max(1, int(oh * scale))
    ws.add_image(xl_img, f"{letter}{row}")
    return min(200.0, max(40.0, xl_img.height * (72.0 / 96.0) + 10.0))


def merge_store(employee: str, store: str, folders: list[Path], output_dir: Path,
                week_mon: date = None, week_sun: date = None):
    print(f"\n  Merge {employee}/{store}: {len(folders)} folders")

    # 收集所有 xlsx
    all_xlsx = []
    for folder in sorted(folders):
        xlsx_files = _find_xlsx_in_folder(folder)
        for xf in xlsx_files:
            all_xlsx.append((folder, xf))
    if not all_xlsx:
        print(f"    skip: no xlsx")
        return

    print(f"    found {len(all_xlsx)} xlsx files")

    # 用最后一个 xlsx 的 header 作为模板（最新格式）
    template_wb = load_workbook(str(all_xlsx[-1][1]))
    template_ws = template_wb.active
    col_count = template_ws.max_column

    # 读取所有行，去重
    seen_no = set()
    # (no_val, source_folder, source_ws, source_row, source_xlsx_path)
    merged_rows = []

    for folder, xlsx_path in all_xlsx:
        print(f"    read: {xlsx_path.name} ({folder.name})")
        try:
            wb = load_workbook(str(xlsx_path))
        except Exception as e:
            print(f"    [error] cannot open {xlsx_path}: {e}")
            continue
        ws = wb.active
        if ws.max_row < 2:
            continue
        col_count = max(col_count, ws.max_column)

        for r in range(2, ws.max_row + 1):
            no_val = ws.cell(r, 1).value
            # 去重：有 No. 的行按 No. 去重，空的（变体子行）总是保留
            if no_val is not None and str(no_val).strip() != "":
                no_key = str(no_val).strip()
                if no_key in seen_no:
                    # 跳过这行，但也跳过它后面的变体子行
                    continue
                seen_no.add(no_key)
            else:
                # 变体子行：检查前一个主行是否被保留
                # 如果前面的主行是重复的（被跳过），则这个子行也跳过
                pass
            merged_rows.append((no_val, folder, ws, r, xlsx_path))

    if not merged_rows:
        print(f"    skip: no data rows")
        return

    # 文件名
    seq_nums = []
    for no_val, _, _, _, _ in merged_rows:
        if no_val is not None and str(no_val).strip() != "":
            try:
                seq_nums.append(int(no_val))
            except (ValueError, TypeError):
                pass
    week_label = ""
    if week_mon and week_sun:
        week_label = f"_{week_mon.strftime('%m%d')}-{week_sun.strftime('%m%d')}"
    if seq_nums:
        min_seq, max_seq = min(seq_nums), max(seq_nums)
        out_name = f"shein_products_{store}{week_label}_{min_seq}-{max_seq}_merged.xlsx"
    else:
        out_name = f"shein_products_{store}{week_label}_merged.xlsx"

    out_path = output_dir / out_name
    print(f"    output: {out_name} ({len(merged_rows)} rows after dedup)")

    # 创建合并 workbook
    new_wb = Workbook()
    new_ws = new_wb.active
    new_ws.title = "Shein Products"

    # 复制 header（从最新模板）
    for c in range(1, col_count + 1):
        src_cell = template_ws.cell(1, c)
        dest_cell = new_ws.cell(1, c, src_cell.value)
        dest_cell.font = copy(src_cell.font)
        dest_cell.fill = copy(src_cell.fill)
        dest_cell.alignment = copy(src_cell.alignment)
        dest_cell.border = copy(src_cell.border)
        letter = get_column_letter(c)
        src_dim = template_ws.column_dimensions.get(letter)
        if src_dim and src_dim.width:
            new_ws.column_dimensions[letter].width = src_dim.width

    new_ws.row_dimensions[1].height = template_ws.row_dimensions[1].height or 24
    new_ws.freeze_panes = "A2"
    new_ws.auto_filter.ref = f"A1:{get_column_letter(col_count)}1"

    # 复制数据行 + 从磁盘插入图片
    dest_row = 2
    img_count = 0
    for no_val, src_folder, src_ws, src_r, xlsx_path in merged_rows:
        # 复制所有单元格值和格式
        for c in range(1, col_count + 1):
            src_cell = src_ws.cell(src_r, c)
            dest_cell = new_ws.cell(dest_row, c, src_cell.value)
            dest_cell.font = copy(src_cell.font)
            if src_cell.fill and src_cell.fill.patternType:
                dest_cell.fill = copy(src_cell.fill)
            dest_cell.alignment = copy(src_cell.alignment)
            dest_cell.border = copy(src_cell.border)
            if src_cell.number_format:
                dest_cell.number_format = src_cell.number_format

        # 行高
        src_height = src_ws.row_dimensions[src_r].height
        if src_height:
            new_ws.row_dimensions[dest_row].height = src_height

        # 图片：从磁盘 seq 文件夹直接插入
        if no_val is not None and str(no_val).strip() != "":
            img_path = _find_first_image(src_folder, no_val)
            if img_path:
                pt = _add_picture(new_ws, dest_row, PICTURE_COL, img_path)
                if pt > 0:
                    img_count += 1
                    # 确保行高够放图片
                    cur_h = new_ws.row_dimensions[dest_row].height or 18
                    new_ws.row_dimensions[dest_row].height = max(cur_h, pt)

        dest_row += 1

    new_wb.save(str(out_path))
    print(f"    done: {dest_row - 2} rows, {img_count} images -> {out_path.name}")


def run_merge(employee_filter: str = "", store_filter: str = "",
              week_date: date = None, all_weeks: bool = False):
    """
    week_date: 指定某一周（取该日期所在周），默认本周。
    all_weeks: True 时忽略周过滤，合并所有日期。
    """
    week_mon, week_sun = _get_week_range(week_date)
    employees = [employee_filter.upper()] if employee_filter else KNOWN_EMPLOYEES

    if not all_weeks:
        print(f"Week: {week_mon} (Mon) ~ {week_sun} (Sun)\n")

    for emp in employees:
        emp_dir = COMPLETED_ROOT / emp
        if not emp_dir.is_dir():
            continue

        store_groups = defaultdict(list)
        for folder in emp_dir.iterdir():
            if not folder.is_dir():
                continue
            if folder.name.startswith("_"):
                continue
            # 跳过已有的 merged xlsx 文件（不是文件夹）
            store = _parse_store_code(folder.name, emp)
            if not store:
                continue

            # 周日期过滤
            if not all_weeks:
                fd = _parse_folder_date(folder.name)
                if fd is None or fd < week_mon or fd > week_sun:
                    continue

            store_groups[store].append(folder)

        if not store_groups:
            continue

        print(f"{'='*50}")
        print(f"Employee: {emp} -- {len(store_groups)} store(s)")

        for store, folders in sorted(store_groups.items()):
            if store_filter and store.upper() != store_filter.upper():
                continue
            if not folders:
                continue
            merge_store(emp, store, folders, emp_dir,
                        week_mon=week_mon if not all_weeks else None,
                        week_sun=week_sun if not all_weeks else None)


def main():
    parser = argparse.ArgumentParser(description="Merge shein_products Excel by employee/store")
    parser.add_argument("employee", nargs="?", default="", help="Employee code (e.g. LUMEI)")
    parser.add_argument("store", nargs="?", default="", help="Store code (e.g. C2)")
    parser.add_argument("--week", default="", help="Date to select week (YYYYMMDD), default=this week")
    parser.add_argument("--all", action="store_true", help="Merge all dates, ignore week filter")
    args = parser.parse_args()

    week_date = None
    if args.week:
        week_date = datetime.strptime(args.week, "%Y%m%d").date()

    run_merge(args.employee, args.store, week_date=week_date, all_weeks=args.all)


if __name__ == "__main__":
    main()
