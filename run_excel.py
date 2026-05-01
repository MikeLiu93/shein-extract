"""
Excel-based pipeline: read pending URLs from .xlsx worksheets (one per store),
scrape them, write results to Listing - completed 2nd/{store}/, and update
Date + Status columns in the source Excel.

Usage:
    python run_excel.py                          # default: Shein Submited Links.xlsx
    python run_excel.py "path/to/file.xlsx"      # specific file

Columns (strict):
    A: Seq       — sequence number (= output folder name). READ ONLY.
    B: Website   — Shein product URL. READ ONLY.
    C: Date      — filled after run: YYYY-MM-DD. WRITE.
    D: Status    — filled after run: Done / Failed / Delisted. WRITE.

Only rows with BOTH Date and Status empty are processed.
"""

import argparse
import logging
import os
import shutil
import sys
import time
import traceback
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook

from shein_scraper import scrape_shein, RateLimitError
from config import SUBMITTED_DIR, OUTPUT_ROOT_2ND as OUTPUT_ROOT, BACKUP_DIR, INPUT_FILENAME

logger = logging.getLogger("run_excel")
DEFAULT_XLSX = INPUT_FILENAME
DEBUG_LOG_DIR = Path(__file__).resolve().parent / "debug_logs"


def setup_logging():
    DEBUG_LOG_DIR.mkdir(parents=True, exist_ok=True)
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    log_path = DEBUG_LOG_DIR / f"excel_{ts}.log"
    fmt = logging.Formatter("%(asctime)s [%(levelname)s] %(message)s")

    root = logging.getLogger("run_excel")
    root.handlers.clear()
    root.setLevel(logging.DEBUG)

    fh = logging.FileHandler(log_path, encoding="utf-8")
    fh.setLevel(logging.DEBUG)
    fh.setFormatter(fmt)
    root.addHandler(fh)

    sh = logging.StreamHandler(sys.stdout)
    sh.setLevel(logging.INFO)
    sh.setFormatter(fmt)
    root.addHandler(sh)

    logger.info("Log: %s", log_path)
    return log_path


def backup_excel(xlsx_path: Path) -> None:
    """Backup Excel to personal Drive before reading, with date suffix."""
    BACKUP_DIR.mkdir(parents=True, exist_ok=True)
    today = datetime.now().strftime("%Y%m%d")
    stem = xlsx_path.stem
    backup_name = f"{stem}_{today}.xlsx"
    dest = BACKUP_DIR / backup_name
    try:
        shutil.copy2(xlsx_path, dest)
        logger.info("Backup: %s → %s", xlsx_path.name, dest)
    except Exception as e:
        logger.warning("Backup failed (continuing anyway): %s", e)


def safe_save(wb, xlsx_path: Path) -> None:
    """Save workbook. If locked by another user, save as copy with '2' suffix."""
    try:
        wb.save(xlsx_path)
    except PermissionError:
        alt = xlsx_path.with_stem(xlsx_path.stem + "2")
        logger.warning("Cannot save to %s (locked), saving to %s", xlsx_path.name, alt.name)
        wb.save(alt)
        logger.info("Saved to alternate: %s", alt.name)


def process_excel(xlsx_path: Path) -> None:
    """Process all worksheets in an Excel file."""
    logger.info("Opening: %s", xlsx_path.name)
    wb = load_workbook(xlsx_path)

    for ws_name in wb.sheetnames:
        ws = wb[ws_name]
        store = ws_name.strip()
        logger.info("Sheet: %s", store)

        # Normalize header: accept "Execute Date" or "Date" in col C
        header_c = str(ws.cell(1, 3).value or "").strip()
        if header_c in ("Execute Date", ""):
            ws.cell(1, 3).value = "Date"

        # Collect pending rows (Date AND Status both empty)
        pending = []
        for row in range(2, ws.max_row + 1):
            seq = ws.cell(row, 1).value
            url = ws.cell(row, 2).value
            date_val = ws.cell(row, 3).value
            status_val = ws.cell(row, 4).value
            if url and seq is not None and not date_val and not status_val:
                pending.append((row, int(seq), str(url).strip()))

        if not pending:
            logger.info("  No pending rows in '%s'", store)
            continue

        logger.info("  %d pending URL(s): seq %s",
                     len(pending), [p[1] for p in pending])

        # Prepare output folder
        store_dir = OUTPUT_ROOT / store
        store_dir.mkdir(parents=True, exist_ok=True)

        # Run scraper
        urls = [p[2] for p in pending]
        seqs = [p[1] for p in pending]
        today = datetime.now().strftime("%Y-%m-%d")
        seq_min, seq_max = min(seqs), max(seqs)
        xlsx_name = f"{store}-{seq_min}-{seq_max}-{today.replace('-', '')}.xlsx"

        # screenshots/ subfolder is created lazily by the scraper itself
        old_cwd = Path.cwd()
        results = None
        try:
            # Google Drive sync may briefly lock new folders
            for _retry in range(5):
                try:
                    os.chdir(store_dir)
                    break
                except PermissionError:
                    time.sleep(2)
            else:
                os.chdir(store_dir)  # final attempt, let it raise
            logger.info("  Scraping %d URLs → %s/%s", len(urls), store, xlsx_name)
            results = scrape_shein(urls, output=xlsx_name, seq_list=seqs)
        except RateLimitError:
            logger.warning("  [限流] Rate limited during '%s'", store)
        except Exception as e:
            logger.exception("  Error processing '%s': %s", store, e)
            try:
                tb = traceback.format_exc()
                (DEBUG_LOG_DIR / "last_traceback.txt").write_text(tb, encoding="utf-8")
            except OSError:
                pass
        finally:
            os.chdir(old_cwd)

        # Update Date + Status based on results
        for row_idx, seq, url in pending:
            seq_folder = store_dir / str(seq)
            has_files = seq_folder.is_dir() and any(seq_folder.iterdir())

            # Determine status from result record
            rec = None
            if results:
                rec = next((r for r in results if r.get("seq_num") == seq), None)

            # A record with empty SKU or [goods_name] title is not a real success
            is_bad_data = (rec and rec.get("status") == "OK"
                           and (not rec.get("sku")
                                or "[goods_name]" in (rec.get("title") or "")))

            if has_files and not is_bad_data:
                ws.cell(row_idx, 3).value = today
                ws.cell(row_idx, 4).value = "Done"
                logger.info("    seq %d → Done", seq)
            elif rec and rec.get("status") == "DELISTED":
                ws.cell(row_idx, 3).value = today
                ws.cell(row_idx, 4).value = "Delisted"
                logger.info("    seq %d → Delisted", seq)
            else:
                detail = rec.get("status", "") if rec else ""
                if is_bad_data:
                    detail = "no data loaded"
                ws.cell(row_idx, 3).value = today
                ws.cell(row_idx, 4).value = "Failed"
                logger.info("    seq %d → Failed %s", seq,
                            f"({detail})" if detail else "")

        # Save after each store (so progress isn't lost)
        safe_save(wb, xlsx_path)
        logger.info("  Saved progress to %s", xlsx_path.name)

    wb.close()
    logger.info("Done: %s", xlsx_path.name)


def main():
    parser = argparse.ArgumentParser(
        description="Excel-based Shein scraper pipeline")
    parser.add_argument("file", nargs="?", default=None,
                        help="Path to .xlsx file (default: Shein Submited Links.xlsx)")
    args = parser.parse_args()

    setup_logging()

    if args.file:
        files = [Path(args.file)]
    else:
        default = SUBMITTED_DIR / DEFAULT_XLSX
        if default.exists():
            files = [default]
        else:
            logger.error("Default file not found: %s", default)
            return

    for f in files:
        logger.info("=" * 60)
        # Backup before reading
        backup_excel(f)
        try:
            process_excel(f)
        except Exception as e:
            logger.exception("Fatal error processing %s: %s", f.name, e)

    logger.info("All done.")


if __name__ == "__main__":
    main(