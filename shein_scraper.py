"""
Shein Product Scraper  (改进版 v2)
=====================
自包含，直接调用 scrape_shein() — 无需手动配置 Chrome。

改进点（相比 2026-03-29 稳定版）：
  1. 智能等待：轮询 goods_sn 出现，最长等 PAGE_LOAD_MAX_WAIT 秒（取代固定9秒睡眠）
  2. 并行图片下载：ThreadPoolExecutor，速度提升 5–10×
  3. goods_imgs JSON 解析：从页面内嵌 JSON 直接获取有序商品主图（更准确）
  4. 页面滚动：下载前滚动商品图集，触发懒加载
  5. 关闭用完的标签页：避免 Chrome 堆积大量标签

INSTALL (once):
    pip install requests websocket-client openpyxl

USAGE:
    from shein_scraper import scrape_shein

    scrape_shein("https://us.shein.com/some-product-p-123456789.html")
    scrape_shein(["url1", "url2"], output="my_results.xlsx")

EXCEL COLUMNS:
  1. No.          行号（变体子行留空）
  2. Date         今天日期
  3. SKU          Shein 内部 goods_sn
  4. Picture      第一张商品图（嵌入）
  5. Price        售价（USD）
  6. Shipping     $0.00（满免运费）或 $2.99
  7. Product URL  原始链接
  8. Store name   店铺名称
  9. Title        商品标题
  10. eBay price  max(Price+Shipping, 22) × 1.4
  11. Variation 1 第一个变体属性（含单值属性）
  12. Variation 2 第二个变体属性（含单值属性）
  13. eBay Title  eBay 上架标题
  * 变体子行的 C 列 (SKU) 填写该变体的 sku_code
"""

import json
import os
import random
import re
import shutil
import subprocess
import tempfile
import threading
import time
from concurrent.futures import ThreadPoolExecutor, as_completed
from pathlib import Path
from urllib.parse import parse_qs, urlencode, urlparse, unquote, urlunparse

import requests
import websocket
from datetime import date
from notify import alert_captcha, alert_signin, alert_generic
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.utils.cell import coordinate_to_tuple


# ── Settings ──────────────────────────────────────────────────────────────────
DEFAULT_SHIPPING_FEE   = 2.99
EBAY_MIN_SUBTOTAL      = float(os.environ.get("SHEIN_EBAY_MIN_SUBTOTAL", "22"))
EBAY_MARKUP            = float(os.environ.get("SHEIN_EBAY_MARKUP", "1.4"))
CDP_PORT               = 9222
PAGE_LOAD_MIN_WAIT     = 3        # 最少等待秒数（让 JS 初始化）
PAGE_LOAD_MAX_WAIT     = 20       # 最多等待秒数（轮询 goods_sn）
PAGE_LOAD_POLL_INTERVAL = 0.5     # 轮询间隔秒数
PAGE_LOAD_RETRIES      = 3        # 遇到错误页面的最大重试次数
RELOAD_PAUSE_SEC       = 2
# Inter-URL pacing (anti-rate-limit). Random jitter + occasional long pauses
# look less robotic than a fixed delay.
INTER_URL_DELAY_MIN    = 4        # 每个商品之间最短间隔（秒）
INTER_URL_DELAY_MAX    = 12       # 每个商品之间最长间隔（秒）
LONG_PAUSE_EVERY       = 18       # 每处理 N 个商品后插一段长歇
LONG_PAUSE_MIN         = 30       # 长歇最短秒数
LONG_PAUSE_MAX         = 90       # 长歇最长秒数
# OOPS retry backoff: Shein 软封禁经常返回 "Oops" 假页面，先退避重试再判 DELISTED
OOPS_RETRY_BACKOFF_MIN = 30
OOPS_RETRY_BACKOFF_MAX = 60
EXTRACTION_TIMEOUT_SEC = 60       # 单个页面提取超时（秒），正常3-5s加载，超过60s视为失败
RATE_LIMIT_CONSECUTIVE  = 3       # 连续失败 N 次视为限流，停止当前批次
PICTURE_MAX_HEIGHT_PX  = 168
KEEP_CHROME_OPEN       = True
PERSISTENT_PROFILE_DIR = os.path.join(os.path.expanduser("~"), "shein-cdp-profile")
OUTPUT_ENCODING        = "utf-8"
MEDIA_FOLDER_PREFIX    = "图片-"
EBAY_LISTING_TXT_NAME  = "eBay上架描述.txt"
IMAGE_DOWNLOAD_WORKERS = 8        # 并行下载线程数
LOW_STOCK_THRESHOLD    = 15       # stock <= 此值标记 [少货]


def _inter_url_pause(i: int, total: int) -> None:
    """Sleep between URLs. Skip after last. Random jitter + occasional long pause."""
    if i >= total:
        return
    delay = random.uniform(INTER_URL_DELAY_MIN, INTER_URL_DELAY_MAX)
    if i > 0 and i % LONG_PAUSE_EVERY == 0:
        long_pause = random.uniform(LONG_PAUSE_MIN, LONG_PAUSE_MAX)
        print(f"  [节奏] 已处理 {i} 条 — 长歇 {long_pause:.0f}s + 间隔 {delay:.1f}s")
        time.sleep(long_pause + delay)
    else:
        print(f"  [节奏] 间隔 {delay:.1f}s")
        time.sleep(delay)


_CHROME_PATHS = [
    r"C:\Program Files\Google\Chrome\Application\chrome.exe",
    r"C:\Program Files (x86)\Google\Chrome\Application\chrome.exe",
    os.path.expandvars(r"%LOCALAPPDATA%\Google\Chrome\Application\chrome.exe"),
    os.path.expandvars(r"%PROGRAMFILES%\Google\Chrome\Application\chrome.exe"),
    os.path.expandvars(r"%PROGRAMFILES(X86)%\Google\Chrome\Application\chrome.exe"),
    "/Applications/Google Chrome.app/Contents/MacOS/Google Chrome",
]


class RateLimitError(Exception):
    """连续多个 URL 失败，判定为 Shein 限流。"""
    pass


# ── JavaScript: 轮询用（只检查 goods_sn 是否出现）────────────────────────────
_JS_POLL = r"""
(function(){
    const scripts = [...document.querySelectorAll('script:not([src])')];
    for (const s of scripts) {
        const m = s.textContent.match(/"goods_sn"\s*:\s*"([^"]+)"/);
        if (m) return m[1];
    }
    // Also check window vars
    try {
        const gs = window?.gbData?.detail?.goods_sn || window?.ProductDetailData?.detail?.goods_sn;
        if (gs) return gs;
    } catch(e) {}
    return null;
})()
"""

# ── JavaScript: 滚动商品图集（触发懒加载）────────────────────────────────────
_JS_SCROLL_GALLERY = r"""
(function(){
    // 找到商品图集容器并滚动
    const gallery = document.querySelector(
        '.product-intro__main, .goods-detail-v3__gallery, [class*="gallery"], [class*="swiper"]'
    );
    if (gallery) {
        gallery.scrollIntoView({behavior: 'instant', block: 'center'});
    }
    // 整体向下滚动一段，触发懒加载
    window.scrollBy(0, 600);
    return true;
})()
"""

# ── JavaScript: 检测登录/验证码拦截 ──────────────────────────────────────────
_JS_DETECT_BLOCK = r"""
(function(){
    var result = {blocked: false, type: null, rect: null};
    var bodyText = (document.body?.innerText || '').toLowerCase();
    var html = document.documentElement?.innerHTML || '';

    // 检测 Sign in / Register 弹窗或页面
    var signinKeywords = ['sign in', 'log in', 'register', 'create account', 'login'];
    var signinEls = document.querySelectorAll(
        '[class*="login"], [class*="signin"], [class*="register"], [id*="login"], [id*="register"]'
    );
    // 检查是否有明显的登录弹窗覆盖了页面
    var modal = document.querySelector(
        '.sui-dialog__wrapper, [class*="modal"][class*="login"], [class*="overlay"][class*="login"]'
    );
    if (modal && modal.offsetHeight > 100) {
        result.blocked = true;
        result.type = 'signin_modal';
        return result;
    }
    // 如果整个页面是登录页
    if (signinKeywords.some(function(k){ return document.title.toLowerCase().includes(k); })) {
        result.blocked = true;
        result.type = 'signin_page';
        return result;
    }

    // 检测验证码/人机验证
    var captchaSelectors = [
        'iframe[src*="captcha"]', 'iframe[src*="challenge"]',
        'iframe[src*="recaptcha"]', 'iframe[src*="hcaptcha"]',
        '[class*="captcha"]', '[id*="captcha"]',
        '[class*="verify"]', '[class*="challenge"]',
        '[class*="slider-verify"]', '[class*="puzzle"]',
        '.geetest_panel', '#geetest', '[class*="geetest"]',
    ];
    for (var i = 0; i < captchaSelectors.length; i++) {
        var el = document.querySelector(captchaSelectors[i]);
        if (el && el.offsetHeight > 30) {
            var r = el.getBoundingClientRect();
            result.blocked = true;
            result.type = 'captcha';
            result.rect = {x: r.left, y: r.top, w: r.width, h: r.height};
            return result;
        }
    }

    // 检测 "Are you a human" / "Verify you are human" 文字
    if (/verify.{0,20}(human|person|real)|are you a (human|robot)|security check/i.test(bodyText)) {
        result.blocked = true;
        result.type = 'captcha_text';
        return result;
    }

    return result;
})()
"""

# ── JavaScript: 尝试关闭登录弹窗 ────────────────────────────────────────────
_JS_DISMISS_SIGNIN = r"""
(function(){
    // 尝试点击各种关闭按钮
    var closeSelectors = [
        '.sui-dialog__headerbtn', '.sui-icon-common__close',
        '[class*="modal"] [class*="close"]', '[class*="dialog"] [class*="close"]',
        '[class*="popup"] [class*="close"]', '[aria-label="Close"]',
        '[class*="login"] [class*="close"]', '.she-close',
        'button[class*="close"]', '.icon-close',
    ];
    for (var i = 0; i < closeSelectors.length; i++) {
        var els = document.querySelectorAll(closeSelectors[i]);
        for (var j = 0; j < els.length; j++) {
            try { els[j].click(); } catch(e) {}
        }
    }
    // 也尝试按 Escape
    document.dispatchEvent(new KeyboardEvent('keydown', {key: 'Escape', keyCode: 27, bubbles: true}));
    return true;
})()
"""


# ── JavaScript: 提取所有商品数据（主提取脚本）────────────────────────────────
_JS = r"""
(function() {
    const out = {};
    const scripts = [...document.querySelectorAll('script:not([src])')];
    const patterns = {
        goods_sn:   /"goods_sn"\s*:\s*"([^"]+)"/,
        goods_id:   /"goods_id"\s*:\s*"?(\d+)"?/,
        store_name: /"store_code"\s*:\s*"[^"]+"\s*,\s*"title"\s*:\s*"([^"]+)"/,
    };
    for (const s of scripts) {
        const t = s.textContent;
        for (const [k, p] of Object.entries(patterns)) {
            if (!out[k]) { const m = t.match(p); if (m) out[k] = m[1]; }
        }
    }
    out.title = (document.querySelector('h1')?.innerText
                 || document.title.split('|')[0])?.trim() || '';
    const priceEl = document.querySelector('.productPrice__main')
        || [...document.querySelectorAll('*')].find(el =>
               /price/i.test(el.className?.toString() || '') &&
               /\$[\d.]+/.test(el.innerText || ''));
    const pm = (priceEl?.innerText || '').match(/([\d.]+)/);
    out.price = pm ? parseFloat(pm[1]) : null;
    const shipEl = document.querySelector('.productShippingNewContent__postage-text')
                || document.querySelector('[class*="postage-text"]');
    const shipText = (shipEl?.innerText || '').trim();
    out.shipping_raw       = shipText;
    const threshM          = shipText.match(/orders?\s*(?:>=|\u2265)\s*\$([\d.]+)/i);
    out.free_threshold     = threshM ? parseFloat(threshM[1]) : null;
    out.unconditional_free = /free\s*shipping/i.test(shipText) && !threshM;
    out.website = 'us.shein.com';
    out.variations = {};

    // -------- Variations (Color / Size / etc.) --------
    function uniq(arr) {
        const seen = new Set();
        const out = [];
        for (const x of arr || []) {
            const v = (x || '').trim();
            if (!v) continue;
            const k = v.toLowerCase();
            if (seen.has(k)) continue;
            seen.add(k);
            out.push(v);
        }
        return out;
    }

    function collectOptions(sec) {
        const nodes = [
            ...sec.querySelectorAll(
                'button,[role="option"],[role="radio"],[aria-label],[data-attr-value],[data-value],[data-sku]'
            ),
            ...sec.querySelectorAll('[class*="attr-item"],[class*="attr-item__val"],[class*="swatch"],[class*="sku"],[class*="size"]'),
        ];
        const vals = [];
        for (const el of nodes) {
            const disabled = el.getAttribute('aria-disabled') === 'true' || el.disabled === true;
            if (disabled) continue;
            const v =
                el.getAttribute('data-attr-value') ||
                el.getAttribute('data-value') ||
                el.getAttribute('aria-label') ||
                (el.innerText || el.textContent || '');
            const t = (v || '').replace(/\s+/g, ' ').trim();
            if (!t) continue;
            if (/^(add to cart|qty|quantity|size guide|sold by|shipping|reviews?)$/i.test(t)) continue;
            if (/^(large image|show more colors|check my size)$/i.test(t)) continue;
            vals.push(t);
        }
        return uniq(vals);
    }

    // A) Size section(s)
    function extractSizes() {
        const labelNode = [...document.querySelectorAll('*')].find(el =>
            (el.childElementCount === 0) && /^(US\s*Size|Size)$/i.test((el.textContent || '').trim())
        );
        const roots = [];
        if (labelNode) {
            let p = labelNode.parentElement;
            for (let i = 0; i < 6 && p; i++) { roots.push(p); p = p.parentElement; }
        }
        roots.push(
            document.querySelector('.product-intro__size'),
            ...document.querySelectorAll('[class*="size"][class*="intro"], [class*="size"] [class*="attr"]')
        );

        const seen = new Set();
        const sizes = [];

        function addSizeText(t) {
            const v = (t || '').replace(/\s+/g, ' ').trim();
            if (!v) return;
            if (!/(US\s*\d|CN\s*\d|EU\s*\d)/i.test(v)) return;
            if (/check my size/i.test(v)) return;
            const k = v.toLowerCase();
            if (seen.has(k)) return;
            seen.add(k);
            sizes.push(v.replace(/\s+\(/g, ' ('));
        }

        for (const r of roots.filter(Boolean)) {
            for (const b of r.querySelectorAll('button,[role="option"],[role="radio"]')) {
                addSizeText(b.innerText || b.textContent || b.getAttribute('aria-label'));
            }
            const txt = (r.innerText || '').replace(/\s+/g, ' ');
            const m = txt.match(/\bUS\s*\d+(?:\.\d+)?\s*\(CN\s*\d+\)\b/gi) || [];
            for (const x of m) addSizeText(x);
        }
        return sizes;
    }

    const allSizes = extractSizes();
    if (allSizes.length) out.variations['US Size'] = allSizes;

    // B) Color / other attribute sections
    const attrSecs = document.querySelectorAll(
        '.product-intro__color, [class*="sales-attr__fold"]:not(.product-intro__size)'
    );
    for (const sec of attrSecs) {
        if (sec.classList.contains('product-intro__size')) continue;
        const hdr = (sec.querySelector('[class*="header"]')
                    || sec.querySelector('[class*="title"]'))
                    ?.innerText?.split('\n')[0]?.trim() || '';
        const ci  = hdr.indexOf(':');
        const lbl = ci > -1 ? hdr.slice(0, ci).trim() : hdr;
        if (!lbl) continue;
        let selected = ci > -1 ? hdr.slice(ci + 1).trim() : '';

        if (!selected) {
            const selectedNode =
                sec.querySelector('[aria-checked="true"]') ||
                sec.querySelector('[aria-selected="true"]') ||
                sec.querySelector('[aria-current="true"]') ||
                sec.querySelector('[class*="selected"],[class*="active"],[class*="current"]');
            selected = (selectedNode?.innerText || selectedNode?.textContent || selectedNode?.getAttribute('aria-label') || '').trim();
        }

        let opts = [...sec.querySelectorAll('[class*="attr-item__val"],[class*="swatch-item"]')]
                        .map(el => el.innerText?.trim()).filter(Boolean);
        if (!opts.length) opts = collectOptions(sec);
        opts = uniq(opts)
            .map(x => x.replace(/\s+/g, ' ').trim())
            .filter(x => x && !/^(large image|show more colors|check my size)$/i.test(x))
            .map(x => x.replace(/^color\s+/i, ''));
        opts = uniq(opts);

        if (selected) {
            selected = selected.replace(/^color\s+/i, '').replace(/\s+/g, ' ').trim();
            if (!/^(large image|show more colors|check my size)$/i.test(selected)) {
                out.variations[lbl] = [selected];
                continue;
            }
        }

        if (opts.length) out.variations[lbl] = opts;
    }

    // -------- goods_imgs JSON 解析（从 "<goods_id>": [{"origin_image":...}] 中提取）--------
    function extractGalleryFromJson() {
        const urls = [];
        const seenUrls = new Set();
        function addUrl(u) {
            if (!u || typeof u !== 'string') return;
            let v = u.trim();
            // 补全协议头
            if (v.startsWith('//')) v = 'https:' + v;
            if (!v || seenUrls.has(v)) return;
            seenUrls.add(v);
            urls.push(v);
        }

        // 用 goods_id 作为键去搜索 [{"origin_image":"..."},...]
        const gid = out.goods_id;
        if (gid) {
            const searchKey = '"' + gid + '"';
            for (const s of scripts) {
                const t = s.textContent || '';
                let searchIdx = 0;
                while (true) {
                    const idx = t.indexOf(searchKey, searchIdx);
                    if (idx < 0) break;
                    // 找冒号，然后找 [
                    const colonIdx = t.indexOf(':', idx + searchKey.length);
                    if (colonIdx < 0 || colonIdx > idx + searchKey.length + 5) {
                        searchIdx = idx + searchKey.length;
                        continue;
                    }
                    let vStart = colonIdx + 1;
                    while (vStart < t.length && t[vStart] === ' ') vStart++;
                    if (t[vStart] === '[') {
                        let depth = 0;
                        let end = -1;
                        for (let j = vStart; j < t.length && j < vStart + 100000; j++) {
                            if (t[j] === '[' || t[j] === '{') depth++;
                            else if (t[j] === ']' || t[j] === '}') {
                                depth--;
                                if (depth === 0) { end = j; break; }
                            }
                        }
                        if (end > 0) {
                            try {
                                const arr = JSON.parse(t.slice(vStart, end + 1));
                                if (Array.isArray(arr) && arr.length > 0 && arr[0].origin_image) {
                                    for (const item of arr) {
                                        addUrl(item.origin_image || '');
                                    }
                                    if (urls.length > 0) return urls;
                                }
                            } catch(e) {}
                        }
                    }
                    searchIdx = idx + searchKey.length;
                }
            }
        }

        // Fallback: 搜索 "detail_image" 数组
        for (const s of scripts) {
            const t = s.textContent || '';
            if (!t.includes('"detail_image"')) continue;
            let searchIdx = 0;
            while (true) {
                const idx = t.indexOf('"detail_image"', searchIdx);
                if (idx < 0) break;
                const colonIdx = t.indexOf(':', idx + 14);
                if (colonIdx < 0) { searchIdx = idx + 14; continue; }
                let vStart = colonIdx + 1;
                while (vStart < t.length && t[vStart] === ' ') vStart++;
                if (t[vStart] === '[') {
                    let depth = 0;
                    let end = -1;
                    for (let j = vStart; j < t.length && j < vStart + 100000; j++) {
                        if (t[j] === '[' || t[j] === '{') depth++;
                        else if (t[j] === ']' || t[j] === '}') {
                            depth--;
                            if (depth === 0) { end = j; break; }
                        }
                    }
                    if (end > 0) {
                        try {
                            const arr = JSON.parse(t.slice(vStart, end + 1));
                            if (Array.isArray(arr) && arr.length > 0) {
                                for (const item of arr) {
                                    addUrl(item.origin_image || item.src || item.url || '');
                                }
                                if (urls.length > 0) return urls;
                            }
                        } catch(e) {}
                    }
                }
                searchIdx = idx + 14;
            }
        }

        return urls;
    }

    const goodsImgUrls = extractGalleryFromJson();
    if (goodsImgUrls.length > 0) {
        out.goods_imgs = goodsImgUrls;
    }

    // -------- Media URLs (images/videos) — 作为 fallback --------
    function absUrl(u) {
        try { return new URL(u, location.href).toString(); } catch (e) { return ""; }
    }
    function dedupe(urls) {
        const out = [];
        const seen = new Set();
        for (const u of urls || []) {
            const v = (u || "").trim();
            if (!v) continue;
            const key = v.split("?")[0];
            if (seen.has(key)) continue;
            seen.add(key);
            out.push(v);
        }
        return out;
    }

    function dedupeWebpPreferLarge(urls) {
        function score(u) {
            let s = 0;
            const m = u.match(/\b(\d{2,4})x(\d{2,4})\b/g);
            if (m) {
                for (const x of m) {
                    const ab = x.split("x");
                    const a = parseInt(ab[0], 10), b = parseInt(ab[1], 10);
                    if (a && b) s = Math.max(s, a * b);
                }
            }
            try {
                const sp = new URL(u, location.href).searchParams;
                const w = parseInt(sp.get("imwidth") || sp.get("width") || sp.get("w") || "0", 10);
                if (w > 0) s = Math.max(s, w * w);
            } catch (e) {}
            return s * 10000 + u.length;
        }
        const byStem = new Map();
        for (const u of urls || []) {
            const v = (u || "").trim();
            if (!v) continue;
            const pathOnly = v.split("?")[0];
            const stem = pathOnly.replace(/\.webp$/i, "");
            const prev = byStem.get(stem);
            if (!prev || score(v) > score(prev)) byStem.set(stem, v);
        }
        return [...byStem.values()];
    }

    function looksLikeSheinWebp(u) {
        const v = (u || "").trim().toLowerCase();
        if (!v || /data:image\//i.test(v)) return false;
        const pathOnly = v.split("?")[0];
        const isWebp =
            /\.webp($|\?)/.test(v) ||
            /format[=,]webp/.test(v) ||
            /[?&]type=webp\b/.test(v);
        if (!isWebp) return false;
        if (/\.(jpe?g|png|gif)($|\?)/.test(pathOnly)) return false;
        if (!/ltwebstatic\.com|sheingroup|img\.shein\.com|\.shein\.com/i.test(v)) return false;
        if (/sprite|icon|logo|avatar|badge|emoji|payment|banner|_ad_|promo-badge/.test(v)) return false;
        if (/\b(50|64|72|80|96|100)x(50|64|72|80|96|100)\b/.test(v)) return false;
        return true;
    }

    function deriveWebpCandidatesFromRaster(raw) {
        const out = [];
        try {
            const u = new URL(raw, location.href);
            const h = u.hostname.toLowerCase();
            if (!/ltwebstatic|sheingroup|img\.shein|\.shein\.com/.test(h)) return out;
            const p = u.pathname;
            if (!/\.(jpe?g|png)$/i.test(p)) return out;
            const base = p.replace(/\.(jpe?g|png)$/i, "");
            const origin = u.origin;
            const q = u.search || "";
            const cand = [
                origin + base + ".webp" + q,
                origin + base + "_format,webp.webp" + q,
            ];
            for (const s of cand) {
                if (looksLikeSheinWebp(s)) out.push(s);
            }
        } catch (e) {}
        return out;
    }

    function upgradeSheinThumbnailPath(u) {
        let s = u || "";
        function bumpWh(prefix, w, h) {
            const wi = parseInt(w, 10), hi = parseInt(h, 10);
            if (!wi || !hi) return prefix + w + "x" + h;
            const nw = Math.max(wi, 1200);
            const nh = Math.max(hi, Math.round((nw * hi) / wi));
            return prefix + nw + "x" + nh;
        }
        s = s.replace(/_square_thumbnail_(\d{3,4})x(\d{3,4})(?=[._/])/gi, (_, w, h) =>
            bumpWh("_square_thumbnail_", w, h)
        );
        s = s.replace(/_thumbnail_(\d{3,4})x(\d{3,4})(?=[._/])/gi, (_, w, h) =>
            bumpWh("_thumbnail_", w, h)
        );
        s = s.replace(/_thumbnail_(\d{3,4})(?=[._/])/gi, (_, n) => {
            const v = parseInt(n, 10);
            if (v >= 900) return `_thumbnail_${v}`;
            return `_thumbnail_${Math.max(v, 1050)}`;
        });
        return s;
    }
    function addImwidthLarge(u) {
        try {
            const x = new URL(u, location.href);
            if (!/ltwebstatic|sheingroup|img\.shein|\.shein\.com/i.test(x.hostname)) return u;
            x.searchParams.set("imwidth", "1340");
            return x.toString();
        } catch (e) {
            return u;
        }
    }
    function prepGalleryWebpUrl(u) {
        return addImwidthLarge(upgradeSheinThumbnailPath((u || "").trim()));
    }

    const imgUrls = [];
    const videoUrls = [];

    function pushImgUrl(raw) {
        const a = absUrl(raw);
        if (!a) return;
        if (/data:image\//i.test(a)) return;
        if (/(sprite|icon|logo)\./i.test(a)) return;
        imgUrls.push(a);
    }

    function pushSrcset(attr) {
        if (!attr) return;
        for (const part of String(attr).split(",")) {
            const u = part.trim().split(/\s+/)[0];
            if (u) pushImgUrl(u);
        }
    }

    for (const img of document.querySelectorAll("img")) {
        pushImgUrl(img.currentSrc || img.src || "");
        pushImgUrl(img.getAttribute("data-src") || "");
        pushImgUrl(img.getAttribute("data-original") || "");
        pushImgUrl(img.getAttribute("data-lazy-src") || "");
        pushSrcset(img.getAttribute("srcset") || img.getAttribute("data-srcset") || "");
    }
    for (const srcEl of document.querySelectorAll("picture source[srcset], picture source[src]")) {
        pushSrcset(srcEl.getAttribute("srcset") || "");
        pushImgUrl(srcEl.getAttribute("src") || "");
    }

    for (const v of document.querySelectorAll("video")) {
        const s = v.currentSrc || v.src || "";
        const a = absUrl(s);
        if (a) videoUrls.push(a);
        for (const srcEl of v.querySelectorAll("source")) {
            const a2 = absUrl(srcEl.src || srcEl.getAttribute("src") || "");
            if (a2) videoUrls.push(a2);
        }
    }
    for (const s of scripts) {
        const t = s.textContent || "";
        const matches = t.match(/https?:\/\/[^"'\s]+?\.(?:mp4|m3u8)(?:\?[^"'\s]*)?/gi) || [];
        for (const m of matches) videoUrls.push(m);
    }

    const og = document.querySelector('meta[property="og:image"]')?.getAttribute("content");
    if (og) imgUrls.push(absUrl(og));

    try {
        const entries = performance.getEntriesByType("resource") || [];
        for (const e of entries) {
            const n = e.name || "";
            if (!/\.webp($|[?#])/i.test(n) && !/format[=,]webp/i.test(n)) continue;
            if (!/ltwebstatic\.com|sheingroup|img\.shein|\.shein\.com/i.test(n)) continue;
            pushImgUrl(n);
        }
    } catch (err) {}

    for (const s of scripts) {
        const t = s.textContent || "";
        if (t.length < 100 || t.length > 3_000_000) continue;
        const ms =
            t.match(/https?:\/\/[^"'\\\s<>]*ltwebstatic\.com[^"'\\\s<>]*\.webp/gi) || [];
        for (const m of ms) {
            if (m.length < 80 || m.length > 2048) continue;
            pushImgUrl(m);
        }
    }

    const derived = [];
    for (const u of imgUrls) {
        for (const d of deriveWebpCandidatesFromRaster(u)) derived.push(d);
    }
    const merged = new Set();
    for (const u of imgUrls.filter(looksLikeSheinWebp)) merged.add(prepGalleryWebpUrl(u));
    for (const u of derived) merged.add(prepGalleryWebpUrl(u));
    let imagesOut = dedupeWebpPreferLarge([...merged]);
    if (!imagesOut.length) {
        const loose = imgUrls.filter(
            (u) => /\.webp($|\?)/i.test((u || "").trim()) && !/sprite|icon|logo|avatar/i.test((u || "").toLowerCase())
        );
        imagesOut = dedupeWebpPreferLarge(loose.map(prepGalleryWebpUrl));
    }

    function urlScoreQuick(u) {
        let s = 0;
        const m = (u || "").match(/\b(\d{2,4})x(\d{2,4})\b/g);
        if (m) {
            for (const x of m) {
                const ab = x.split("x");
                const a = parseInt(ab[0], 10), b = parseInt(ab[1], 10);
                if (a && b) s = Math.max(s, a * b);
            }
        }
        try {
            const sp = new URL(u, location.href).searchParams;
            const w = parseInt(sp.get("imwidth") || sp.get("width") || sp.get("w") || "0", 10);
            if (w > 0) s = Math.max(s, w * w);
        } catch (e) {}
        return s * 10000 + (u || "").length;
    }
    imagesOut = dedupeWebpPreferLarge(imagesOut.map(prepGalleryWebpUrl));
    imagesOut = imagesOut.filter((u) => {
        const m = (u || "").match(/_thumbnail_(\d{3,4})(?=[._/])/i);
        if (!m) return true;
        return parseInt(m[1], 10) >= 600;
    });
    imagesOut.sort((a, b) => urlScoreQuick(b) - urlScoreQuick(a));
    if (imagesOut.length > 120) imagesOut = imagesOut.slice(0, 120);

    if (!imagesOut.length && merged.size) {
        imagesOut = dedupeWebpPreferLarge([...merged]);
        imagesOut.sort((a, b) => urlScoreQuick(b) - urlScoreQuick(a));
        if (imagesOut.length > 120) imagesOut = imagesOut.slice(0, 120);
    }

    out.media = {
        images: imagesOut,
        videos: dedupe(videoUrls),
    };

    // -------- Per-variant pricing (from window.gbRawData) --------
    function extractSkuPrices() {
        try {
            const gb = window.gbRawData;
            if (!gb || !gb.modules || !gb.modules.saleAttr) return [];
            const multi = gb.modules.saleAttr.multiLevelSaleAttribute;
            if (!multi || !multi.sku_list) return [];
            return multi.sku_list.map(function(sku) {
                var attrs = {};
                var arr = sku.sku_sale_attr || [];
                for (var i = 0; i < arr.length; i++) {
                    var a = arr[i];
                    attrs[a.attr_name || String(a.attr_id)] = a.attr_value_name || '';
                }
                var sp = null, rp = null;
                try { sp = sku.priceInfo.salePrice.amount; } catch(e) {}
                if (sp === null) try { sp = sku.price.salePrice.amount; } catch(e) {}
                try { rp = sku.priceInfo.retailPrice.amount; } catch(e) {}
                if (rp === null) try { rp = sku.price.retailPrice.amount; } catch(e) {}
                return {
                    sku_code: sku.sku_code || '',
                    attrs: attrs,
                    sale_price: sp,
                    retail_price: rp,
                    stock: sku.stock || 0
                };
            });
        } catch(e) { return []; }
    }
    out.sku_prices = extractSkuPrices();

    // -------- mainSaleAttribute: 颜色变体对应的独立 goods_sn --------
    function extractMainSaleAttr() {
        try {
            var gb = window.gbRawData;
            if (!gb || !gb.modules || !gb.modules.saleAttr) return [];
            var main = gb.modules.saleAttr.mainSaleAttribute;
            if (!main || !main.info) return [];
            return main.info.map(function(item) {
                return {
                    goods_id: item.goods_id || '',
                    goods_sn: item.goods_sn || '',
                    attr_value_name: item.attr_value || item.attr_value_name || '',
                    attr_name: item.attr_name || main.attr_name || '',
                    is_current: !!item.isMainGood
                };
            });
        } catch(e) { return []; }
    }
    out.main_sale_attrs = extractMainSaleAttr();

    return out;
})()
"""


# ── Chrome launcher ───────────────────────────────────────────────────────────

def _find_chrome():
    for path in _CHROME_PATHS:
        if os.path.exists(path):
            return path
    raise RuntimeError(
        "Chrome not found. Searched:\n" + "\n".join(f"  {p}" for p in _CHROME_PATHS)
    )


def _launch_chrome(port=CDP_PORT, profile_dir=None):
    chrome = _find_chrome()
    if profile_dir is None:
        profile_dir = tempfile.mkdtemp(prefix="shein_chrome_")
    else:
        os.makedirs(profile_dir, exist_ok=True)

    cmd = [
        chrome,
        f"--remote-debugging-port={port}",
        "--remote-allow-origins=*",
        f"--user-data-dir={profile_dir}",
        "--no-first-run",
        "--no-default-browser-check",
        "--disable-extensions",
        "--disable-popup-blocking",
        "--disable-blink-features=AutomationControlled",
        "--window-size=1280,900",
        "--new-window",
        "about:blank",
    ]

    proc = subprocess.Popen(
        cmd,
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
    )

    for attempt in range(30):
        try:
            resp = requests.get(f"http://localhost:{port}/json", timeout=2)
            tabs = resp.json()
            if any(t.get("type") == "page" for t in tabs):
                print(f"  Chrome ready on port {port}")
                return proc, profile_dir
        except Exception:
            pass
        time.sleep(0.5)

    proc.kill()
    raise RuntimeError(
        f"Chrome launched (PID {proc.pid}) but port {port} never became available.\n"
        "Try closing other Chrome windows and running again."
    )


def _cdp_available(port=CDP_PORT) -> bool:
    try:
        resp = requests.get(f"http://localhost:{port}/json", timeout=2)
        return resp.status_code == 200
    except Exception:
        return False


def _ensure_chrome(port=CDP_PORT):
    if _cdp_available(port):
        print(f"  Reusing existing Chrome on port {port}")
        return None, None, False
    proc, profile_dir = _launch_chrome(port, profile_dir=PERSISTENT_PROFILE_DIR)
    return proc, profile_dir, True


# ── CDP WebSocket ─────────────────────────────────────────────────────────────

def _cdp_once(ws_url, method, params=None, timeout=30):
    result = {}
    done   = threading.Event()

    def on_open(ws):
        ws.send(json.dumps({"id": 1, "method": method, "params": params or {}}))

    def on_message(ws, msg):
        d = json.loads(msg)
        if d.get("id") == 1:
            result["data"] = d.get("result", {})
            done.set()
            ws.close()

    def on_error(ws, err):
        if "closed" not in str(err).lower():
            result["error"] = str(err)
        done.set()

    ws = websocket.WebSocketApp(
        ws_url,
        on_open=on_open,
        on_message=on_message,
        on_error=on_error,
    )
    threading.Thread(target=ws.run_forever, daemon=True).start()
    done.wait(timeout=timeout)

    if "error" in result:
        raise RuntimeError(f"CDP error: {result['error']}")
    return result.get("data", {})


def _get_ws_url(port=CDP_PORT):
    tabs = requests.get(f"http://localhost:{port}/json", timeout=5).json()
    pages = [t for t in tabs if t.get("type") == "page"]
    if not pages:
        requests.get(f"http://localhost:{port}/json/new", timeout=5)
        time.sleep(1)
        tabs = requests.get(f"http://localhost:{port}/json", timeout=5).json()
        pages = [t for t in tabs if t.get("type") == "page"]
    return pages[0]["webSocketDebuggerUrl"]


def _new_tab(port=CDP_PORT, url=None):
    """Create a new tab. If url given, opens directly at that URL — equivalent
    to user typing the URL into a fresh tab's address bar (no Referer)."""
    endpoint = f"http://localhost:{port}/json/new"
    if url:
        from urllib.parse import quote
        endpoint += "?" + quote(url, safe="")
    tab = requests.put(endpoint, timeout=5).json()
    return tab.get("id"), tab.get("webSocketDebuggerUrl")


def _close_tab(port, tab_id):
    """关闭指定标签页（用完即关，避免 Chrome 堆积）。"""
    try:
        requests.get(f"http://localhost:{port}/json/close/{tab_id}", timeout=5)
    except Exception:
        pass


def _ws_url_for_id(port, tab_id):
    tabs = requests.get(f"http://localhost:{port}/json", timeout=5).json()
    for t in tabs:
        if t.get("id") == tab_id:
            return t.get("webSocketDebuggerUrl")
    raise RuntimeError(f"Could not find tab id {tab_id} on port {port}")


def _run_js(ws_url, js):
    res = _cdp_once(ws_url, "Runtime.evaluate", {
        "expression":    js,
        "returnByValue": True,
        "awaitPromise":  False,
        "userGesture":   True,
    })
    if res.get("exceptionDetails"):
        raise RuntimeError(f"JS error: {res['exceptionDetails']}")
    return res.get("result", {}).get("value")


# ── 智能等待：轮询 goods_sn（取代固定 sleep）────────────────────────────────

def _wait_for_page_ready(port, tab_id, min_wait=PAGE_LOAD_MIN_WAIT,
                          max_wait=PAGE_LOAD_MAX_WAIT, poll=PAGE_LOAD_POLL_INTERVAL):
    """
    先等 min_wait 秒，然后每 poll 秒检查一次 goods_sn 是否出现在页面中，
    最多等 max_wait 秒。返回找到的 goods_sn 或 None（超时）。
    """
    time.sleep(min_wait)
    ws_url = _ws_url_for_id(port, tab_id)
    elapsed = min_wait
    while elapsed < max_wait:
        try:
            val = _run_js(ws_url, _JS_POLL)
            if val:
                print(f"  页面就绪（等待 {elapsed:.1f}s，goods_sn={val}）")
                return val, ws_url
        except Exception:
            pass
        time.sleep(poll)
        elapsed += poll
        ws_url = _ws_url_for_id(port, tab_id)  # 刷新以防 WS URL 变化
    print(f"  等待超时（{max_wait}s），继续尝试提取...")
    return None, _ws_url_for_id(port, tab_id)


_SHEIN_REFERRER = "https://us.shein.com/"

# Anti-detection JS injected before every page load
_JS_ANTI_DETECT = """
Object.defineProperty(navigator, 'webdriver', {get: () => undefined});
if (!window.chrome) { window.chrome = {}; }
if (!window.chrome.runtime) { window.chrome.runtime = {}; }
"""

# Whether session warmup (homepage visit) has been done in this process
_SESSION_WARMED: bool = False


def _ensure_shein_session(port):
    """
    On first call, ensure cookies/session are warm by visiting Shein homepage.
    Idempotent: subsequent calls are no-ops within the same process.
    """
    global _SESSION_WARMED
    if _SESSION_WARMED:
        return

    tabs = requests.get(f"http://localhost:{port}/json", timeout=5).json()
    pages = [t for t in tabs if t.get("type") == "page"]
    shein_pages = [t for t in pages if "shein.com" in (t.get("url") or "")]

    # If a Shein page already exists from a previous run, treat session as warm
    if shein_pages:
        _SESSION_WARMED = True
        return

    # Otherwise, do warmup in any existing tab (or create one)
    if pages:
        tab = pages[0]
    else:
        tab = requests.put(f"http://localhost:{port}/json/new", timeout=5).json()
    ws_url = tab.get("webSocketDebuggerUrl") or _ws_url_for_id(port, tab["id"])

    try:
        _cdp_once(ws_url, "Page.addScriptToEvaluateOnNewDocument",
                  {"source": _JS_ANTI_DETECT})
    except Exception:
        pass

    print("  [导航] 预热：先访问 Shein 首页建立 session...")
    _cdp_once(ws_url, "Page.navigate", {
        "url": _SHEIN_REFERRER,
        "referrer": "",
        "transitionType": "typed",
    })
    time.sleep(5)
    _SESSION_WARMED = True


def _navigate_and_wait(port, url):
    """
    Open a fresh tab DIRECTLY at the target URL via PUT /json/new?<url>.
    Single-step equivalent of "user opens new tab + pastes URL + Enter".
    No Page.navigate, no Referer, no chance of empty-referrer weirdness.
    The scrape loop closes the tab in its finally clause after extraction.
    Returns (ws_url, tab_id).
    """
    _ensure_shein_session(port)

    print(f"  [导航] 新 tab @ URL → {url[:90]}")
    tab_id, ws_url = _new_tab(port, url=url)
    if not tab_id or not ws_url:
        raise RuntimeError("Failed to create new tab via /json/new?<url>")

    time.sleep(1)
    for _ in range(10):
        try:
            ws_url = _ws_url_for_id(port, tab_id)
            break
        except Exception:
            time.sleep(0.5)
    else:
        raise RuntimeError(f"Tab {tab_id} lost after creation")

    _, ws_url = _wait_for_page_ready(port, tab_id)

    # Diagnostic: confirm Referer behavior (should be empty for typed nav)
    try:
        actual_ref = _run_js(ws_url, "document.referrer")
        print(f"  [导航] document.referrer = '{actual_ref}'")
    except Exception:
        pass

    return _ws_url_for_id(port, tab_id), tab_id


def _reload_tab_and_wait(port, tab_id):
    ws_url = _ws_url_for_id(port, tab_id)
    _cdp_once(ws_url, "Page.reload", {"ignoreCache": True})
    _, ws_url = _wait_for_page_ready(port, tab_id)
    return _ws_url_for_id(port, tab_id)


def _take_screenshot(ws_url, save_path: str, clip_rect=None) -> bool:
    """通过 CDP 截图并保存为 PNG。clip_rect={x,y,w,h} 时只截该矩形（带 10px padding）。"""
    try:
        import base64
        params = {"format": "png"}
        if clip_rect:
            pad = 10
            x = max(0.0, float(clip_rect.get("x", 0)) - pad)
            y = max(0.0, float(clip_rect.get("y", 0)) - pad)
            w = float(clip_rect.get("w", 0)) + pad * 2
            h = float(clip_rect.get("h", 0)) + pad * 2
            if w > 0 and h > 0:
                params["clip"] = {"x": x, "y": y, "width": w, "height": h, "scale": 1}
        res = _cdp_once(ws_url, "Page.captureScreenshot", params, timeout=15)
        data_b64 = res.get("data")
        if data_b64:
            Path(save_path).write_bytes(base64.b64decode(data_b64))
            return True
    except Exception as e:
        print(f"  [截图失败] {e}")
    return False


def _screenshots_dir(base_dir) -> Path:
    """Ensure screenshots/ subfolder under base_dir; return its path."""
    d = Path(base_dir) / "screenshots"
    d.mkdir(parents=True, exist_ok=True)
    return d


def _check_and_handle_block(port, tab_id, url, base_dir, max_wait=300):
    """
    检测登录弹窗/验证码拦截。
    - 登录弹窗：自动尝试关闭
    - 验证码：立即截图 + 发邮件通知，等待人工解决（不再自动绕过）
    - 超时未解决：返回 False
    返回 True 表示已解决或无拦截，False 表示仍被拦截。
    """
    ws_url = _ws_url_for_id(port, tab_id)
    try:
        result = _run_js(ws_url, _JS_DETECT_BLOCK)
    except Exception:
        return True  # 检测失败，继续正常流程

    if not isinstance(result, dict) or not result.get("blocked"):
        return True  # 没有被拦截

    block_type = result.get("type", "unknown")
    print(f"  [拦截检测] 类型: {block_type}")

    # ── 登录弹窗：尝试自动关闭 ──
    if "signin" in block_type:
        print("  [拦截处理] 尝试关闭登录弹窗...")
        for attempt in range(3):
            try:
                ws_url = _ws_url_for_id(port, tab_id)
                _run_js(ws_url, _JS_DISMISS_SIGNIN)
                time.sleep(1.5)
                ws_url = _ws_url_for_id(port, tab_id)
                check = _run_js(ws_url, _JS_DETECT_BLOCK)
                if not isinstance(check, dict) or not check.get("blocked"):
                    print("  [拦截处理] 登录弹窗已关闭")
                    return True
            except Exception:
                pass
        # 3 次尝试失败，截图 + 通知
        print("  [拦截处理] 登录弹窗关闭失败，发送通知...")
        ss_path = str(_screenshots_dir(base_dir) / f"_block_{time.strftime('%Y%m%d_%H%M%S')}.png")
        try:
            ws_url = _ws_url_for_id(port, tab_id)
            _take_screenshot(ws_url, ss_path)
        except Exception:
            ss_path = None
        alert_signin(url, screenshot_path=ss_path)
        return False

    # ── 验证码：立即邮件通知，等待人工解决（不自动绕过） ──
    if "captcha" in block_type:
        print(f"  [拦截处理] 检测到人机验证，发送邮件通知，等待人工（最多 {max_wait}s）...")
        ss_path = str(_screenshots_dir(base_dir) / f"_captcha_{time.strftime('%Y%m%d_%H%M%S')}.png")
        try:
            ws_url = _ws_url_for_id(port, tab_id)
            _take_screenshot(ws_url, ss_path, clip_rect=result.get("rect"))
        except Exception:
            ss_path = None
        alert_captcha(url, screenshot_path=ss_path)

        elapsed = 0
        poll = 15
        while elapsed < max_wait:
            time.sleep(poll)
            elapsed += poll
            try:
                ws_url = _ws_url_for_id(port, tab_id)
                check = _run_js(ws_url, _JS_DETECT_BLOCK)
                if not isinstance(check, dict) or not check.get("blocked"):
                    print(f"  [拦截处理] 验证码已通过（等了 {elapsed}s）")
                    return True
            except Exception:
                pass
            print(f"  [拦截处理] 仍在等待验证码... ({elapsed}/{max_wait}s)")

        print("  [拦截处理] 验证码等待超时")
        return False

    return True


def _shein_page_needs_retry(data) -> bool:
    if not isinstance(data, dict):
        return True
    t = (data.get("title") or "").strip().lower()
    err_markers = (
        "bad gateway", "502 bad", "503", "504",
        "gateway time-out", "gateway timeout",
        "service unavailable", "error 502", "error 503",
        "nginx", "cloudflare",
    )
    if any(m in t for m in err_markers):
        return True
    if not (data.get("goods_sn") or str(data.get("goods_id") or "").strip()):
        if t in ("bad gateway", "502", "503", "504", "error"):
            return True
        if "error" in t and len(t) < 40 and not data.get("price"):
            return True
    return False


# ── Shipping & pricing ────────────────────────────────────────────────────────

def _calc_shipping(data):
    price     = data.get("price") or 0.0
    threshold = data.get("free_threshold")
    uncon     = data.get("unconditional_free", False)
    if uncon:                  return 0.0
    if threshold is not None:  return 0.0 if price >= threshold else DEFAULT_SHIPPING_FEE
    return DEFAULT_SHIPPING_FEE


def _ebay_listing_price(price: float, shipping: float) -> float:
    base = float(price or 0) + float(shipping or 0)
    return round(max(base, EBAY_MIN_SUBTOTAL) * EBAY_MARKUP, 2)


def _split_variations_for_excel(variations: dict, sku_prices: list = None,
                                shipping: float = 0.0) -> tuple[str, str]:
    """
    Variation 1: 非尺寸属性（颜色等）
    Variation 2: 尺寸属性 + 每个尺寸的价格（如有不同价格）
    """
    if not variations or not isinstance(variations, dict):
        return "", ""

    sku_prices = sku_prices or []

    def fmt_line(key: str) -> str:
        v = variations[key]
        vals = v if isinstance(v, list) else [v]
        return f"{key}: {', '.join(map(str, vals))}"

    keys = list(variations.keys())
    size_keys = [k for k in keys if re.search(r"\bsize\b", k, re.I)]
    size_set = set(size_keys)
    non_size_keys = [k for k in keys if k not in size_set]

    # 构建 size → price 映射
    size_price_map = {}
    if sku_prices:
        for sp in sku_prices:
            attrs = sp.get("attrs") or {}
            sale = sp.get("sale_price")
            if sale is None:
                continue
            # 找到 size 属性
            for ak, av in attrs.items():
                if re.search(r"\bsize\b", ak, re.I) and av:
                    size_price_map[av] = float(sale)
                    break

    # 检查是否所有 size 价格都一样
    unique_prices = set(size_price_map.values())
    prices_vary = len(unique_prices) > 1

    def fmt_size_line(key: str) -> str:
        v = variations[key]
        vals = v if isinstance(v, list) else [v]
        if prices_vary and size_price_map:
            parts = []
            for sv in vals:
                p = size_price_map.get(sv)
                if p is not None:
                    ebay_p = _ebay_listing_price(p, shipping)
                    parts.append(f"{sv}: ${p:.2f}→eBay${ebay_p:.2f}")
                else:
                    parts.append(str(sv))
            return f"{key}:\n" + "\n".join(f"  {p}" for p in parts)
        return f"{key}: {', '.join(map(str, vals))}"

    v1_parts = [fmt_line(k) for k in non_size_keys] if non_size_keys else []
    v2_parts = [fmt_size_line(k) for k in size_keys] if size_keys else []

    if size_keys and non_size_keys:
        return "\n".join(v1_parts), "\n".join(v2_parts)
    if size_keys:
        return "\n".join(v2_parts), ""
    if len(keys) == 1:
        return fmt_line(keys[0]), ""
    return "\n".join(fmt_line(k) for k in keys), ""


# ── Image helpers ─────────────────────────────────────────────────────────────

def _first_product_image_path(folder) -> "Path | None":
    if folder is None or not Path(folder).is_dir():
        return None
    p = Path(folder) / "img_001.webp"
    if p.is_file():
        return p
    imgs = sorted(Path(folder).glob("img_*.*"))
    return imgs[0] if imgs else None


def _remove_images_at_cell(ws, row: int, col: int) -> None:
    imgs = getattr(ws, "_images", None) or []
    if not imgs:
        return
    keep: list = []
    for im in imgs:
        a = im.anchor
        r = c = None
        if isinstance(a, str):
            try:
                r, c = coordinate_to_tuple(a.upper())
            except Exception:
                r = c = None
        elif a is not None and hasattr(a, "_from"):
            r, c = a._from.row + 1, a._from.col + 1
        if r == row and c == col:
            continue
        keep.append(im)
    ws._images = keep


def _picture_column_inner_width_px(ws, col: int) -> int:
    letter = get_column_letter(col)
    dim = ws.column_dimensions.get(letter)
    wchars = float(dim.width) if dim and dim.width else float(_COLS[col - 1][1])
    return max(24, int((wchars * 7.0 + 5.0) * 0.90))


def _add_picture_to_cell(ws, row: int, col: int, image_path: "Path") -> int:
    if not Path(image_path).is_file():
        return 0
    try:
        xl_img = XLImage(str(image_path))
    except Exception:
        return 0
    ow, oh = max(1, int(xl_img.width)), max(1, int(xl_img.height))
    max_w = _picture_column_inner_width_px(ws, col)
    max_h = PICTURE_MAX_HEIGHT_PX
    scale = min(max_w / ow, max_h / oh, 1.0)
    nw, nh = max(1, int(ow * scale)), max(1, int(oh * scale))
    xl_img.width = nw
    xl_img.height = nh
    col_letter = get_column_letter(col)
    _remove_images_at_cell(ws, row, col)
    ws.add_image(xl_img, f"{col_letter}{row}")
    return min(200.0, max(40.0, nh * (72.0 / 96.0) + 10.0))


# ── Excel writer ──────────────────────────────────────────────────────────────

_HDR = PatternFill("solid", start_color="1A73E8", end_color="1A73E8")
_ALT = PatternFill("solid", start_color="EEF4FF", end_color="EEF4FF")
_ERR = PatternFill("solid", start_color="FFE0E0", end_color="FFE0E0")
_BRD = Border(**{s: Side(style="thin", color="D0D0D0")
                 for s in ("left", "right", "top", "bottom")})
_COLS = [
    ("No.", 6),
    ("Date", 12),
    ("SKU", 22),
    ("Picture", 14),
    ("Price", 11),
    ("Shipping", 11),
    ("Product URL", 52),
    ("Store name", 20),
    ("Title", 48),
    ("eBay price", 12),
    ("Variation 1", 36),
    ("Variation 2", 36),
    ("eBay Title", 48),
    ("Stock", 16),
]
PICTURE_COL = 4

def _expand_records(records: list) -> list:
    """
    将含多个 sku_prices 的记录展开为多行：
    - 每行对应一个 SKU 变体，V1/V2 显示该变体的具体属性值
    - 所有变体属性都会显示（包括只有单一值的属性如 Size: Double）
    - 第一行为完整信息行，后续行为变体子行（No.留空）
    - 每行都带有各自的 Variant SKU
    """
    expanded = []
    for rec in records:
        sku_prices = rec.get("sku_prices") or []
        variations = rec.get("variations") or {}
        shipping = rec.get("shipping") or 0.0
        main_sale_attrs = rec.get("main_sale_attrs") or []

        # 构建 颜色值 → goods_sn 映射（mainSaleAttribute 中每个颜色是独立商品）
        color_to_goodssn = {}
        for msa in main_sale_attrs:
            val = (msa.get("attr_value_name") or "").strip()
            gsn = (msa.get("goods_sn") or "").strip()
            if val and gsn:
                color_to_goodssn[val.lower()] = gsn

        if len(sku_prices) <= 1:
            # 单变体或无变体，不展开，记录库存
            if sku_prices:
                stk = int(sku_prices[0].get("stock") or 0)
                if stk == 0:
                    rec["_stock"] = "缺货"
                elif 0 < stk <= LOW_STOCK_THRESHOLD:
                    rec["_stock"] = f"少货 only {stk} left"
                else:
                    rec["_stock"] = str(stk)
            expanded.append(rec)
            continue

        # 从 sku_prices 收集有序属性键
        attr_keys_ordered = []
        seen_keys_lower = set()
        for sp in sku_prices:
            for ak in (sp.get("attrs") or {}):
                lk = ak.lower()
                if lk not in seen_keys_lower:
                    seen_keys_lower.add(lk)
                    attr_keys_ordered.append(ak)

        # 补充 variations dict 中有但 sku_prices 中没有的键
        for pk in variations:
            if pk.lower() not in seen_keys_lower:
                seen_keys_lower.add(pk.lower())
                attr_keys_ordered.append(pk)

        if not attr_keys_ordered:
            expanded.append(rec)
            continue

        for i, sp in enumerate(sku_prices):
            attrs = sp.get("attrs") or {}
            sale = sp.get("sale_price")
            stock = int(sp.get("stock") or 0)

            # 变体 SKU：用 mainSaleAttribute 的 goods_sn（按颜色匹配）
            # 仅当颜色对应独立商品时才有 goods_sn，否则留空（父行已有 SKU）
            variant_sku = ""
            for av in attrs.values():
                gsn = color_to_goodssn.get((av or "").strip().lower(), "")
                if gsn:
                    variant_sku = gsn
                    break

            # 为每个属性键获取该 SKU 的具体值
            v_texts = []
            for k in attr_keys_ordered:
                val = ""
                # 先从 sku attrs 匹配（精确 + 忽略大小写）
                val = attrs.get(k, "")
                if not val:
                    for ak, av in attrs.items():
                        if ak.lower() == k.lower():
                            val = av
                            break
                # 兜底：从 variations dict 取值（单值属性等）
                if not val:
                    v = variations.get(k)
                    if not v:
                        for vk, vv in variations.items():
                            if vk.lower() == k.lower():
                                v = vv
                                break
                    if isinstance(v, list) and len(v) == 1:
                        val = str(v[0])
                    elif isinstance(v, list) and v:
                        val = ", ".join(map(str, v))
                    elif v:
                        val = str(v)
                if val:
                    v_texts.append(f"{k}: {val}")

            # 库存标注加到最后一个变体文本上
            if v_texts:
                if stock == 0:
                    v_texts[-1] += " [缺货]"
                elif 0 < stock <= LOW_STOCK_THRESHOLD:
                    v_texts[-1] += f" [少货 only {stock} left]"

            v1 = v_texts[0] if len(v_texts) > 0 else ""
            v2 = "\n".join(v_texts[1:]) if len(v_texts) > 1 else ""

            if sale is not None:
                sale_f = float(sale)
                ebay_p = _ebay_listing_price(sale_f, shipping)
            else:
                sale_f = rec.get("price") or 0.0
                ebay_p = rec.get("ebay_price") or 0.0

            # 库存标注
            if stock == 0:
                stock_label = "缺货"
            elif 0 < stock <= LOW_STOCK_THRESHOLD:
                stock_label = f"少货 only {stock} left"
            else:
                stock_label = str(stock)

            if i == 0:
                # 第一行：完整信息
                row = dict(rec)
                row["price"] = sale_f
                row["ebay_price"] = ebay_p
                row["_v1_text"] = v1
                row["_v2_text"] = v2
                row["_expanded"] = True
                row["_variant_sku"] = variant_sku
                row["_stock"] = stock_label
            else:
                # 后续行：变体子行（No.留空）
                row = {
                    "is_variant_row": True,
                    "_expanded": True,
                    "seq_num": rec.get("seq_num"),
                    "sku": "",  # 不重复写 goods_sn
                    "status": rec.get("status", "OK"),
                    "price": sale_f,
                    "shipping": shipping,
                    "ebay_price": ebay_p,
                    "_v1_text": v1,
                    "_v2_text": v2,
                    "_variant_sku": variant_sku,
                    "_stock": stock_label,
                }

            expanded.append(row)

    return expanded


def _cell(ws, r, c, v, bold=False, fmt=None, wrap=False, fill=None):
    x = ws.cell(r, c, v)
    x.font      = Font(name="Arial", size=10, bold=bold)
    x.alignment = Alignment(vertical="center", wrap_text=wrap)
    x.border    = _BRD
    if fmt:  x.number_format = fmt
    if fill: x.fill = fill
    return x

def _save_excel(records, path):
    path = str(path)
    today = date.today()

    def ensure_header(ws):
        for ci, (h, w) in enumerate(_COLS, 1):
            c = ws.cell(1, ci, h)
            c.font = Font(name="Arial", bold=True, color="FFFFFF", size=11)
            c.fill = _HDR
            c.alignment = Alignment(horizontal="center", vertical="center")
            c.border = _BRD
            ws.column_dimensions[get_column_letter(ci)].width = w
        ws.row_dimensions[1].height = 24
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = f"A1:{get_column_letter(len(_COLS))}1"

    def renumber_rows(ws, recs) -> None:
        # 用 rec 里的 seq_num（如有），否则 fallback 到行号
        # 变体子行（SKU 为空且 URL 也为空）No. 留空
        sku_seq = {}
        url_seq = {}
        for rc in recs:
            sn = str(rc.get("sku") or "").strip()
            if sn:
                sku_seq[sn] = rc.get("seq_num")
            url = str(rc.get("url") or "").strip()
            if url and rc.get("seq_num") is not None:
                url_seq[url] = rc.get("seq_num")
        for r in range(2, ws.max_row + 1):
            sku_val = str(ws.cell(r, 3).value or "").strip()
            url_val = str(ws.cell(r, 7).value or "").strip()
            if sku_val:
                seq = sku_seq.get(sku_val)
                ws.cell(r, 1, seq if seq is not None else r - 1)
            elif url_val:
                # Failed record: no SKU but has URL → match by URL
                seq = url_seq.get(url_val)
                ws.cell(r, 1, seq if seq is not None else "")
            else:
                ws.cell(r, 1, "")  # 变体子行：No. 留空

    def existing_skus(ws) -> set[str]:
        skus = set()
        for r in range(2, ws.max_row + 1):
            v = ws.cell(r, 3).value
            if v:
                skus.add(str(v).strip())
        return skus

    def sku_row_map(ws) -> dict[str, int]:
        m = {}
        for r in range(2, ws.max_row + 1):
            v = ws.cell(r, 3).value
            if v:
                m[str(v).strip()] = r
        return m

    def url_row_map(ws) -> dict[str, int]:
        """URL (col 7) → row，用于 retry 时按 URL 覆盖错误占位行。"""
        m = {}
        for r in range(2, ws.max_row + 1):
            v = ws.cell(r, 7).value
            if v:
                m[str(v).strip()] = r
        return m

    if os.path.exists(path):
        wb = load_workbook(path)
        ws = wb.active
        if ws.max_row >= 1 and (ws.cell(1, 1).value or "") != "No.":
            print(
                "  [Excel] Old column layout detected. "
                "Clearing sheet and applying new headers."
            )
            ws.delete_rows(1, ws.max_row)
            ensure_header(ws)
        elif ws.max_row < 1:
            ensure_header(ws)
        skus = existing_skus(ws)
        row_map = sku_row_map(ws)
        u_row_map = url_row_map(ws)
        start_row = ws.max_row + 1
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = "Shein Products"
        ensure_header(ws)
        skus = set()
        row_map = {}
        u_row_map = {}
        start_row = 2

    ri = start_row
    for rec in records:
        sku = str(rec.get("sku") or "").strip()
        rec_url = str(rec.get("url") or "").strip()
        if sku and sku in row_map:
            target_row = row_map[sku]
        elif rec_url and rec_url in u_row_map:
            # Retry 模式：通过 URL 匹配覆盖错误占位行
            target_row = u_row_map[rec_url]
        else:
            target_row = ri
            if sku:
                skus.add(sku)
                row_map[sku] = target_row
            ri += 1

        p = rec.get("price") or 0.0
        s = rec.get("shipping") or 0.0
        ok = rec.get("status") == "OK"
        eb = rec.get("ebay_price")
        if eb is None and ok:
            eb = _ebay_listing_price(p, s)

        is_variant = rec.get("is_variant_row", False)
        is_expanded = rec.get("_expanded", False)

        # 变体展开模式用 _v1_text/_v2_text，否则用旧逻辑
        if is_expanded:
            v1 = rec.get("_v1_text", "")
            v2 = rec.get("_v2_text", "")
        else:
            v1, v2 = _split_variations_for_excel(
                rec.get("variations") or {},
                sku_prices=rec.get("sku_prices"),
                shipping=s,
            )

        bg = _ERR if rec.get("status") != "OK" else (_ALT if target_row % 2 == 0 else None)

        variant_sku = rec.get("_variant_sku", "")
        ebay_title = rec.get("ebay_title", "")
        stock_label = rec.get("_stock", "")

        if is_variant:
            # 变体子行：C 列填 variant SKU，加 price/shipping/eBay price/V1/V2/Stock
            _cell(ws, target_row, 3, variant_sku, fill=bg)
            _cell(ws, target_row, 5, p, fmt='"$"#,##0.00', fill=bg)
            _cell(ws, target_row, 6, s, fmt='"$"#,##0.00', fill=bg)
            _cell(
                ws, target_row, 10,
                eb if ok and eb is not None else "",
                fmt='"$"#,##0.00' if ok and eb is not None else None,
                bold=True, fill=bg,
            )
            _cell(ws, target_row, 11, v1, wrap=True, fill=bg)
            _cell(ws, target_row, 12, v2, wrap=True, fill=bg)
            _cell(ws, target_row, 14, stock_label, fill=bg)
            ws.row_dimensions[target_row].height = 18
        else:
            # 完整行：所有列
            _cell(ws, target_row, 2, today, fmt="YYYY-MM-DD", fill=bg)
            _cell(ws, target_row, 3, sku, fill=bg)
            _cell(ws, target_row, PICTURE_COL, "", fill=bg)
            _cell(ws, target_row, 5, p, fmt='"$"#,##0.00', fill=bg)
            _cell(ws, target_row, 6, s, fmt='"$"#,##0.00', fill=bg)
            _cell(ws, target_row, 7, rec.get("url") or rec.get("website", ""), wrap=True, fill=bg)
            _cell(ws, target_row, 8, rec.get("store_name", ""), fill=bg)
            _cell(ws, target_row, 9, rec.get("title", ""), wrap=True, fill=bg)
            _cell(
                ws, target_row, 10,
                eb if ok and eb is not None else "",
                fmt='"$"#,##0.00' if ok and eb is not None else None,
                bold=True, fill=bg,
            )
            _cell(ws, target_row, 11, v1, wrap=True, fill=bg)
            _cell(ws, target_row, 12, v2, wrap=True, fill=bg)
            _cell(ws, target_row, 13, ebay_title, wrap=True, fill=bg)
            _cell(ws, target_row, 14, stock_label, fill=bg)
            pic_path = rec.get("first_image_path")
            has_pic = bool(pic_path and Path(str(pic_path)).is_file())
            line_count = max(1, v1.count("\n") + 1, v2.count("\n") + 1)
            base_text_pt = max(18.0, 15.0 * line_count)
            img_pt = 0.0
            if has_pic:
                img_pt = float(
                    _add_picture_to_cell(ws, target_row, PICTURE_COL, Path(str(pic_path)))
                )
            ws.row_dimensions[target_row].height = max(18, int(max(base_text_pt, img_pt)))

        # Write seq_num to No. column immediately (don't rely on renumber_rows)
        seq_num = rec.get("seq_num")
        if seq_num is not None and not is_variant:
            ws.cell(target_row, 1, seq_num)

    renumber_rows(ws, records)
    wb.save(path)


# ── AI-powered eBay title generation ──────────────────────────────────────────

_ANTHROPIC_API_KEY: str | None = None

# Set by worker before calling scrape_shein — controls title style per employee/store.
_TITLE_STYLE: str = ""

_STYLE_INSTRUCTIONS: dict[str, str] = {
    "NA":    "Style: lead with product features and specifications.",
    "TT":    "Style: lead with use case and target audience.",
    "YAN":   "Style: lead with size/dimensions and material.",
    "ZQW":   "Style: lead with product category and key benefit.",
    "LUMEI": "Style: lead with color/design and functionality.",
}


def _get_api_key() -> str | None:
    global _ANTHROPIC_API_KEY
    if _ANTHROPIC_API_KEY is not None:
        return _ANTHROPIC_API_KEY or None
    # Try .env in project dir
    env_path = Path(__file__).resolve().parent / ".env"
    if env_path.exists():
        for line in env_path.read_text(encoding="utf-8").splitlines():
            if line.startswith("ANTHROPIC_API_KEY="):
                _ANTHROPIC_API_KEY = line.split("=", 1)[1].strip()
                return _ANTHROPIC_API_KEY
    # Fallback to environment variable
    _ANTHROPIC_API_KEY = os.environ.get("ANTHROPIC_API_KEY", "")
    return _ANTHROPIC_API_KEY or None


_EBAY_TITLE_PROMPT = """\
Rewrite this SHEIN product title into an eBay listing title (max 80 characters).

Rules:
1. REMOVE brand names (usually the first 1-2 words, e.g. "UMAY", "EastVita", "Pempet", "NIKE PRO", "SheGlam"). Common nouns like "Dog", "Kids" are NOT brands.
2. KEEP all measurements/units with punctuation: 29"-45", 0.25LB, 24"/30"/36", 2", 100ml, 3.5oz.
3. KEEP structural punctuation: inch marks ", ranges -, slashes /, parentheses ().
4. Use the FULL 80 characters as much as possible — prioritize keeping more keywords.
5. ALWAYS start with "NEW " and end with " FREE SHIPPING" if the total is ≤ 80 chars. If too long, try just "NEW " prefix. Drop tags only as a last resort.
6. Do NOT invent information. Only use words from the original title.
7. Remove redundant repeated words (e.g. "Dog Cage Dog Kennel" → "Dog Cage Kennel").

Original: {title}
{style_instruction}
Reply with ONLY the title, no quotes, no explanation."""


def _make_ebay_title_ai(original_title: str) -> str | None:
    """Call Claude Haiku to generate an eBay title. Returns None on failure."""
    api_key = _get_api_key()
    if not api_key:
        return None
    try:
        style_inst = _STYLE_INSTRUCTIONS.get(_TITLE_STYLE, "")
        resp = requests.post(
            "https://api.anthropic.com/v1/messages",
            headers={
                "x-api-key": api_key,
                "anthropic-version": "2023-06-01",
                "content-type": "application/json",
            },
            json={
                "model": "claude-haiku-4-5-20251001",
                "max_tokens": 120,
                "temperature": 0.5,
                "messages": [{"role": "user",
                              "content": _EBAY_TITLE_PROMPT.format(
                                  title=original_title,
                                  style_instruction=style_inst)}],
            },
            timeout=15,
        )
        if resp.status_code != 200:
            return None
        text = resp.json().get("content", [{}])[0].get("text", "").strip()
        # Strip wrapping quotes if present
        if len(text) >= 2 and text[0] == text[-1] and text[0] in ('"', "'"):
            text = text[1:-1].strip()
        # Fix truncated FREE SHIPPING (e.g. "FREE SH", "FREE SHIP")
        m = re.search(r'\s+FREE\s+\S*$', text)
        if m and 'FREE SHIPPING' not in text:
            text = text[:m.start()].rstrip(' ,;-')
        # Sanity: must be non-empty and ≤ 80 chars
        if not text or len(text) > 80:
            return None
        return text
    except Exception:
        return None


# ── Fallback rule-based eBay title ───────────────────────────────────────────

_STOPWORDS = {
    "women", "womens", "men", "mens", "for", "and", "the", "a", "an", "with", "to",
    "of", "in", "on", "by", "from", "shop", "online", "fashion"
}

# Acronyms / units that look like a brand prefix but must be kept.
_BRAND_WHITELIST = {
    "DIY", "USA", "USB", "LED", "PCS", "OZ", "ML", "LB", "LBS", "CM", "MM", "KG",
    "FT", "UV", "UK", "EU", "US", "3D", "4K", "HD", "XL", "XXL", "LCD", "TV",
    "DC", "AC", "RGB", "PVC", "ABS", "BPA", "OEM", "PU", "TPU",
}


def _clean_ws(s: str) -> str:
    return re.sub(r"\s+", " ", (s or "")).strip()


def _strip_brand_prefix(title: str) -> str:
    """Drop leading brand-name tokens (up to first 2).

    Rules:
      - All-caps alpha token, length 2..12 → drop (consecutive, max 2).
      - First-position PascalCase token (initial cap + ≥1 internal cap, alpha, len ≥4) → drop one.
      - Whitelist (unit/acronym) stops the scan.
    """
    parts = title.split()
    drop = 0
    for i, w in enumerate(parts[:2]):
        core = w.strip(",.;:()[]{}!?\"")
        if not core or core in _BRAND_WHITELIST:
            break
        if core.isalpha() and core.isupper() and 2 <= len(core) <= 12:
            drop += 1
            continue
        if i == 0 and core.isalpha() and len(core) >= 4 and core[0].isupper() \
                and any(c.isupper() for c in core[1:]):
            drop += 1
            break
        break
    return " ".join(parts[drop:])


_ARTICLE_RE = re.compile(r'\b(?:the|a|an)\b', re.IGNORECASE)
_NOISE_PHRASE_RE = re.compile(
    r',?\s*\b(?:Shop\s+Online|Free\s+Shipping|Limited\s+Time|Best\s+Seller'
    r'|Hot\s+Sale|New\s+Arrival|SHEIN|SheIn)\b',
    re.IGNORECASE,
)
_SEGMENT_SPLIT_RE = re.compile(r'\s*[,;]\s+|\s+[-–—]+\s+')


def _make_ebay_title(original_title: str, variations: dict, max_len: int = 80) -> str:
    """
    Generate an eBay title (<=80 chars) that preserves the original structure,
    punctuation (inches, ranges, units), and meaning.
    """
    t = _clean_ws(original_title)
    if not t:
        return ""
    t = _strip_brand_prefix(t)

    # 1) Light cleanup: remove articles and marketing noise phrases
    t = _NOISE_PHRASE_RE.sub('', t)
    t = _ARTICLE_RE.sub('', t)
    t = re.sub(r'\(\s*\)', '', t)          # empty parens left after removal
    t = re.sub(r'^\s*[,;]\s*', '', t)      # leading separator
    t = _clean_ws(t)

    # 2) Extract color from variations
    color = ""
    if isinstance(variations, dict):
        for k in variations.keys():
            if "color" in (k or "").lower():
                vals = variations.get(k) or []
                if isinstance(vals, list) and vals:
                    color = _clean_ws(str(vals[0]))
                break

    # 3) Split into segments (comma / dash / semicolon delimited)
    segments = [s.strip() for s in _SEGMENT_SPLIT_RE.split(t) if s.strip()]
    if not segments:
        return ""

    # 4) Join segments front-to-back, fitting within max_len
    core = segments[0]
    for seg in segments[1:]:
        candidate = f"{core}, {seg}"
        if len(candidate) <= max_len:
            core = candidate
        else:
            remaining = max_len - len(core) - 2  # ", "
            if remaining > 10:
                partial = seg[:remaining + 1]
                idx = partial.rfind(' ')
                if idx > remaining // 2:
                    partial = partial[:idx]
                else:
                    partial = partial[:remaining]
                core = f"{core}, {partial.rstrip(' ,;-')}"
            break

    # 5) Append color if not already present
    if color and color.lower() not in core.lower():
        with_color = f"{core} {color}"
        if len(with_color) <= max_len:
            core = with_color

    # 6) Add NEW / FREE SHIPPING tags when space allows
    tag_free = " FREE SHIPPING"
    tag_new = "NEW "

    full = f"{tag_new}{core}{tag_free}"
    if len(full) <= max_len:
        return full
    mid = f"{core}{tag_free}"
    if len(mid) <= max_len:
        return mid
    mid2 = f"{tag_new}{core}"
    if len(mid2) <= max_len:
        return mid2
    if len(core) <= max_len:
        return core

    # Truncate at word boundary
    truncated = core[:max_len + 1]
    idx = truncated.rfind(' ')
    if idx > max_len // 2:
        truncated = truncated[:idx]
    else:
        truncated = truncated[:max_len]
    return truncated.rstrip(' ,-;')


def _safe_filename(s: str) -> str:
    s = _clean_ws(s)
    s = re.sub(r'[<>:"/\\|?*\x00-\x1f]', "_", s)
    return s[:120] if len(s) > 120 else s


def _write_ebay_listing_txt(rec: dict, media_folder) -> "Path | None":
    sku = rec.get("sku") or ""
    if not sku or media_folder is None or not Path(media_folder).is_dir():
        return None

    original_title = _clean_ws(rec.get("original_title", rec.get("title", "")))
    ebay_price = rec.get("ebay_price", "")
    variations = rec.get("variations") or {}
    sku_prices = rec.get("sku_prices") or []
    ebay_title = rec.get("ebay_title") or _make_ebay_title(original_title, variations)

    if isinstance(ebay_price, (int, float)):
        ep_line = f"eBay价格: ${float(ebay_price):.2f}"
    else:
        ep_line = f"eBay价格: {ebay_price}"

    lines = [
        f"希音原标题: {original_title}",
        f"eBay标题: {ebay_title}",
        ep_line,
    ]

    # 变体信息（含每个 size 的价格）
    if isinstance(variations, dict) and variations:
        lines.append("")
        lines.append("变体:")
        for k, v in variations.items():
            if isinstance(v, list):
                lines.append(f"- {k}: {', '.join(map(str, v))}")
            else:
                lines.append(f"- {k}: {v}")

    # 每个 SKU 变体的价格明细
    if sku_prices:
        lines.append("")
        lines.append("各变体价格:")
        for sp in sku_prices:
            attrs = sp.get("attrs") or {}
            sale = sp.get("sale_price")
            stock = int(sp.get("stock") or 0)
            if sale is None:
                continue
            sale_f = float(sale)
            shipping = rec.get("shipping") or 0.0
            ebay_var = _ebay_listing_price(sale_f, shipping)
            attr_str = ", ".join(f"{k}: {v}" for k, v in attrs.items() if v)
            if stock == 0:
                stock_note = " [缺货]"
            elif 0 < stock <= LOW_STOCK_THRESHOLD:
                stock_note = f" [少货 only {stock} left]"
            else:
                stock_note = ""
            lines.append(f"  {attr_str} → eBay${ebay_var:.2f}{stock_note}")

    out_path = Path(media_folder) / EBAY_LISTING_TXT_NAME
    out_path.write_text("\n".join(lines) + "\n", encoding=OUTPUT_ENCODING)
    return out_path


# ── Media downloader ──────────────────────────────────────────────────────────

_DL_HEADERS = {
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36",
    "Referer": "https://us.shein.com/",
}


def _guess_ext(url: str, content_type: "str | None") -> str:
    path = urlparse(url).path
    ext = Path(unquote(path)).suffix.lower()
    if ext and len(ext) <= 6:
        return ext
    if content_type:
        ct = content_type.split(";")[0].strip().lower()
        return {
            "image/jpeg": ".jpg",
            "image/jpg": ".jpg",
            "image/png": ".png",
            "image/webp": ".webp",
            "image/gif": ".gif",
            "video/mp4": ".mp4",
            "application/vnd.apple.mpegurl": ".m3u8",
            "application/x-mpegurl": ".m3u8",
        }.get(ct, "")
    return ""


def _download_one(
    url: str,
    dest: "Path",
    timeout: int = 45,
    headers: "dict | None" = None,
    retries: int = 3,
) -> bool:
    hdr = {**_DL_HEADERS, **(headers or {})}
    for attempt in range(retries):
        try:
            with requests.get(url, headers=hdr, stream=True, timeout=timeout) as r:
                r.raise_for_status()
                ct = (r.headers.get("content-type") or "").lower()
                if ct.startswith("text/") or "text/html" in ct:
                    return False

                ext = _guess_ext(url, ct)
                final = dest.with_suffix(ext or dest.suffix)
                tmp = final.with_suffix((final.suffix or "") + ".part")
                with open(tmp, "wb") as f:
                    for chunk in r.iter_content(chunk_size=1024 * 256):
                        if chunk:
                            f.write(chunk)

                size = tmp.stat().st_size
                is_image = ct.startswith("image/")
                is_video = ct.startswith("video/") or any(
                    (final.suffix or "").lower() == x for x in (".mp4", ".m3u8")
                )
                if is_image and size < 20_000:
                    tmp.unlink(missing_ok=True)
                    return False
                if is_video and size < 150_000:
                    tmp.unlink(missing_ok=True)
                    return False

                tmp.replace(final)
            return True
        except Exception:
            if attempt < retries - 1:
                time.sleep(1.0 + attempt * 0.5)
    return False


def _is_shein_webp_asset_url(u: str) -> bool:
    if not isinstance(u, str) or not u.startswith("http"):
        return False
    ul = u.lower()
    if not any(
        x in ul
        for x in ("ltwebstatic.com", "sheingroup", "img.shein.com", ".shein.com/", "shein.com/")
    ):
        return False
    if any(
        x in ul
        for x in ("sprite", "icon", "logo", "avatar", "badge", "emoji", "payment", "banner", "_ad_", "promo-badge")
    ):
        return False
    path_only = ul.split("?")[0]
    if path_only.endswith((".jpg", ".jpeg", ".png", ".gif")):
        return False
    is_webp = (
        path_only.endswith(".webp")
        or ".webp?" in ul
        or "format=webp" in ul
        or "format,webp" in ul
        or re.search(r"(?:^|[?&])type=webp(?:$|&)", ul) is not None
    )
    if not is_webp:
        return False
    if re.search(r"\b(50|64|72|80|96|100)x(50|64|72|80|96|100)\b", ul):
        return False
    return True


def _webp_url_size_score(u: str) -> int:
    ul = (u or "").lower()
    s = 0
    for m in re.finditer(r"\b(\d{2,4})x(\d{2,4})\b", ul):
        a, b = int(m.group(1)), int(m.group(2))
        s = max(s, a * b)
    try:
        qs = parse_qs(urlparse(u).query)
        for key in ("imwidth", "width", "w"):
            if key in qs and qs[key]:
                w = int(qs[key][0])
                s = max(s, w * w)
                break
    except (ValueError, TypeError):
        pass
    return s * 10_000 + len(u)


_SHEIN_THUMB_W_RE = re.compile(r"_thumbnail_(\d{3,4})(?=[._/])", re.I)
_SHEIN_THUMB_WH_RE = re.compile(
    r"_(square_thumbnail|thumbnail)_(\d{3,4})x(\d{3,4})(?=[._/])", re.I,
)


def _shein_upgrade_webp_to_large(u: str) -> str:
    try:
        p = urlparse(u)
        path = p.path

        def _bump_thumb_wh(m: re.Match) -> str:
            kind, w, h = m.group(1), int(m.group(2)), int(m.group(3))
            if w < 1 or h < 1:
                return m.group(0)
            nw = max(w, 1200)
            nh = max(h, int(round(nw * h / w)))
            return f"_{kind}_{nw}x{nh}"

        def _bump_thumb(m: re.Match) -> str:
            n = int(m.group(1))
            if n >= 900:
                return m.group(0)
            return f"_thumbnail_{max(n, 1050)}"

        path = _SHEIN_THUMB_WH_RE.sub(_bump_thumb_wh, path)
        path = _SHEIN_THUMB_W_RE.sub(_bump_thumb, path)
        q = parse_qs(p.query, keep_blank_values=True)
        host = (p.netloc or "").lower()
        if "ltwebstatic" in host or "shein" in host or "sheingroup" in host:
            q["imwidth"] = ["1340"]
        pairs = [(k, v) for k, vs in q.items() for v in vs]
        new_q = urlencode(pairs) if pairs else ""
        return urlunparse((p.scheme, p.netloc, path, p.params, new_q, p.fragment))
    except Exception:
        return u


def _shein_webp_passes_large_rule(u: str) -> bool:
    ul = (u or "").lower()
    m = _SHEIN_THUMB_W_RE.search(ul)
    if m:
        return int(m.group(1)) >= 600
    edge = 0
    for mm in re.finditer(r"\b(\d{2,4})x(\d{2,4})\b", ul):
        a, b = int(mm.group(1)), int(mm.group(2))
        edge = max(edge, min(a, b))
    try:
        qs = parse_qs(urlparse(u).query)
        for key in ("imwidth", "width", "w"):
            if key in qs and qs[key]:
                edge = max(edge, int(qs[key][0]))
                break
    except (ValueError, TypeError, IndexError):
        pass
    if edge == 0:
        return True
    return edge >= 480


def _download_media(rec: dict, base_dir: "Path", seq_num: "int | None" = None) -> "Path | None":
    """
    改进版媒体下载：
    1. 优先使用 goods_imgs（从 JSON 解析的有序主图）
    2. 并行下载（ThreadPoolExecutor）
    3. 文件夹以 seq_num 命名（如 "26"），无 seq 时回退到 SKU
    """
    sku = rec.get("sku") or ""
    media = rec.get("media") or {}
    goods_imgs = rec.get("goods_imgs") or []
    if not sku or not isinstance(media, dict):
        return None

    if seq_num is not None:
        folder = Path(base_dir) / str(seq_num)
    else:
        folder = Path(base_dir) / f"{MEDIA_FOLDER_PREFIX}{_safe_filename(sku)}"
    folder.mkdir(parents=True, exist_ok=True)

    try:
        for p in folder.iterdir():
            if p.is_file():
                p.unlink(missing_ok=True)
    except Exception:
        pass

    images = media.get("images") or []
    videos = media.get("videos") or []
    if not isinstance(images, list):
        images = []
    if not isinstance(videos, list):
        videos = []

    page_url = (rec.get("url") or "").strip()
    dl_headers = {}
    if page_url:
        dl_headers["Referer"] = page_url

    # ── 优先：goods_imgs 中的图片（有序，直接来自 JSON，无需过滤）──
    priority_urls = []
    if goods_imgs:
        for u in goods_imgs:
            if not isinstance(u, str) or not u.startswith("http"):
                continue
            u2 = _shein_upgrade_webp_to_large(u)
            priority_urls.append(u2)
        print(f"  [图片] goods_imgs 解析到 {len(priority_urls)} 张主图")

    # ── 备选：DOM 扫描图片（过滤 + 升级）──
    fallback_urls = []
    for u in images:
        if not isinstance(u, str) or not _is_shein_webp_asset_url(u):
            continue
        u2 = _shein_upgrade_webp_to_large(u)
        if _shein_webp_passes_large_rule(u2):
            fallback_urls.append(u2)
    fallback_urls.sort(key=_webp_url_size_score, reverse=True)
    fallback_urls = fallback_urls[:120]

    # 当 goods_imgs 有结果时，只用它（准确的商品主图），不混入 DOM 图片
    # 仅当 goods_imgs 为空时才用 DOM fallback
    if priority_urls:
        img_urls = priority_urls
    else:
        img_urls = fallback_urls
        print(f"  [图片] goods_imgs 为空，使用 DOM fallback ({len(fallback_urls)} 张)")

    # ── 并行下载图片 ──
    downloaded_images: list[Path] = []
    download_lock = threading.Lock()

    def dl_img(idx_url):
        idx, u = idx_url
        name = folder / f"_dl_img_{idx:04d}"
        if _download_one(u, name, timeout=55, headers=dl_headers, retries=3):
            candidates = [p for p in folder.glob(name.name + ".*") if p.suffix != ".part"]
            if candidates:
                best = max(candidates, key=lambda p: p.stat().st_size)
                with download_lock:
                    downloaded_images.append((idx, best))
                return True
        return False

    with ThreadPoolExecutor(max_workers=IMAGE_DOWNLOAD_WORKERS) as executor:
        list(executor.map(dl_img, enumerate(img_urls)))

    # 按原始顺序排列（priority 图片在前保持顺序）
    downloaded_images.sort(key=lambda x: x[0])
    ordered_paths = [p for _, p in downloaded_images]

    # goods_imgs 来自 JSON 解析，是精确匹配，放宽文件大小限制
    use_priority = bool(priority_urls)
    MIN_WEBP_BYTES = 30_000 if use_priority else 80_000
    MIN_WEBP_FALLBACK = 10_000 if use_priority else 35_000

    webp_files = [
        p for p in ordered_paths if p.suffix.lower() == ".webp" and p.is_file()
    ]

    large_w = [p for p in webp_files if p.stat().st_size >= MIN_WEBP_BYTES]
    if not large_w:
        large_w = [p for p in webp_files if p.stat().st_size >= MIN_WEBP_FALLBACK]

    keep_set = set(large_w)

    if not keep_set and webp_files:
        tiny = [p for p in webp_files if p.stat().st_size >= 5_000]
        tiny.sort(key=lambda p: p.stat().st_size, reverse=True)
        if tiny:
            keep_set = set(tiny[:25])
        else:
            best = max(webp_files, key=lambda p: p.stat().st_size)
            if best.stat().st_size >= 3_000:
                keep_set = {best}

    # 保留 keep_set，按下载顺序排
    kept_ordered = [p for p in ordered_paths if p in keep_set]
    # 删除其余临时文件
    for p in ordered_paths:
        if p not in keep_set:
            try:
                p.unlink(missing_ok=True)
            except Exception:
                pass

    for i, p in enumerate(kept_ordered, start=1):
        newp = folder / f"img_{i:03d}.webp"
        if p.resolve() != newp.resolve():
            try:
                if newp.exists():
                    newp.unlink(missing_ok=True)
                p.replace(newp)
            except Exception:
                try:
                    shutil.copy2(p, newp)
                    p.unlink(missing_ok=True)
                except Exception:
                    pass

    # ── 视频下载（串行，通常只有 1–2 个）──
    mp4_urls = [u for u in videos if isinstance(u, str) and ".mp4" in u.lower()]
    downloaded_videos: list[Path] = []
    for idx, u in enumerate(mp4_urls):
        name = folder / f"_dl_vid_{idx:04d}"
        if _download_one(u, name, timeout=120, headers=dl_headers, retries=3):
            candidates = [p for p in folder.glob(name.name + ".*") if p.suffix != ".part"]
            if candidates:
                downloaded_videos.append(max(candidates, key=lambda p: p.stat().st_size))

    downloaded_videos.sort(key=lambda p: p.stat().st_size, reverse=True)
    for i, p in enumerate(downloaded_videos, start=1):
        final = folder / f"video_{i:03d}.mp4"
        if p.resolve() != final.resolve():
            try:
                if final.exists():
                    final.unlink(missing_ok=True)
                p.replace(final)
            except Exception:
                try:
                    shutil.copy2(p, final)
                    p.unlink(missing_ok=True)
                except Exception:
                    pass

    return folder


# ── Public function ───────────────────────────────────────────────────────────

def scrape_shein(urls, output="shein_products.xlsx", start_seq=1, seq_list=None):
    """
    抓取一个或多个 Shein 商品 URL，保存到 Excel。

    Parameters
    ----------
    urls      : str | list[str]
    output    : str               输出 .xlsx 文件名
    start_seq : int               Excel 第一列序号起始值
    seq_list  : list[int] | None  每个 URL 对应的序号（用于 retry，覆盖 start_seq）
    """
    if isinstance(urls, str):
        urls = [urls]

    print("Launching Chrome...")
    chrome_proc, profile_dir, launched_new = _ensure_chrome()
    if launched_new:
        print(f"Chrome ready (PID {chrome_proc.pid})  —  scraping {len(urls)} URL(s)\n")
    else:
        print(f"Chrome ready (existing session)  —  scraping {len(urls)} URL(s)\n")

    records = []
    _consecutive_fails = 0
    _rate_limited = False
    try:
        try:
            import state_tracker as _st
        except Exception:
            _st = None
        for i, url in enumerate(urls, 1):
            print(f"[{i}/{len(urls)}] {url[:80]}...")
            if _st is not None:
                try:
                    _st.url_progress(done=i - 1, total=len(urls), current_url=url)
                except Exception:
                    pass
            rec = {
                "url": url,
                "status": "OK",
                "seq_num": seq_list[i - 1] if seq_list else start_seq + (i - 1),
            }
            tab_id = None
            _url_start_time = time.monotonic()
            try:
                print("  navigating...")
                ws_url, tab_id = _navigate_and_wait(CDP_PORT, url)

                # 检测登录/验证码拦截
                base_dir = Path.cwd()
                if not _check_and_handle_block(CDP_PORT, tab_id, url, base_dir):
                    rec["status"] = "BLOCKED"
                    print("  [跳过] 页面被拦截，无法提取")
                    records.append(rec)
                    _inter_url_pause(i, len(urls))
                    continue
                # 重置提取超时计时器：拦截处理（尤其验证码）可能耗时数分钟
                _url_start_time = time.monotonic()

                # 检测商品下架/404（Oops 页面）或数据未加载（[goods_name]）
                _OOPS_DETECT_JS = """
                    (function() {
                        if (document.body && (
                            document.body.innerText.includes('Oops') ||
                            document.querySelector('.page-not-found, .error-page, [class*="not-found"]')
                        )) return 'OOPS';
                        if (document.title && document.title.includes('[goods_name]'))
                            return 'NO_DATA';
                        return 'OK';
                    })()
                """
                try:
                    _page_check = _run_js(ws_url, _OOPS_DETECT_JS)

                    # OOPS 可能是真下架，也可能是 bot 软封禁。先退避重试一次再判定。
                    if _page_check == "OOPS":
                        _backoff = random.uniform(OOPS_RETRY_BACKOFF_MIN, OOPS_RETRY_BACKOFF_MAX)
                        print(f"  [OOPS] 商品页显示 Oops — 退避 {_backoff:.0f}s 后重试...")
                        time.sleep(_backoff)
                        try:
                            ws_url = _reload_tab_and_wait(CDP_PORT, tab_id)
                            _page_check = _run_js(ws_url, _OOPS_DETECT_JS)
                        except Exception:
                            pass
                        if _page_check == "OOPS":
                            rec["status"] = "DELISTED"
                            print("  [跳过] 重试后仍 Oops，判定真下架")
                            records.append(rec)
                            _consecutive_fails = 0
                            _inter_url_pause(i, len(urls))
                            continue
                        print(f"  [OOPS] 重试成功 (state={_page_check})，继续提取")
                        # 重置计时器（退避耗时不计入提取超时）
                        _url_start_time = time.monotonic()

                    if _page_check == "NO_DATA":
                        rec["status"] = "NO_DATA"
                        print("  [跳过] 页面数据未加载 ([goods_name])")
                        records.append(rec)
                        _inter_url_pause(i, len(urls))
                        continue
                except Exception:
                    pass

                # 滚动触发懒加载，稍等片刻
                try:
                    _run_js(ws_url, _JS_SCROLL_GALLERY)
                    time.sleep(1.0)
                    ws_url = _ws_url_for_id(CDP_PORT, tab_id)
                except Exception:
                    pass

                data = None
                for attempt in range(PAGE_LOAD_RETRIES):
                    # 超时保护：如果单个页面总耗时超过 EXTRACTION_TIMEOUT_SEC
                    _elapsed = time.monotonic() - _url_start_time
                    if _elapsed > EXTRACTION_TIMEOUT_SEC:
                        print(f"  [超时] 页面处理已超过 {EXTRACTION_TIMEOUT_SEC}s，可能被隐形拦截")
                        # 二次验证码检测
                        _recheck = None
                        try:
                            ws_url = _ws_url_for_id(CDP_PORT, tab_id)
                            _recheck = _run_js(ws_url, _JS_DETECT_BLOCK)
                        except Exception:
                            pass
                        # 无论是否检测到验证码，都截图通知
                        ss_path = str(_screenshots_dir(base_dir) / f"_timeout_{time.strftime('%Y%m%d_%H%M%S')}.png")
                        try:
                            ws_url = _ws_url_for_id(CDP_PORT, tab_id)
                            _take_screenshot(ws_url, ss_path)
                        except Exception:
                            ss_path = None
                        _block_info = ""
                        if isinstance(_recheck, dict) and _recheck.get("blocked"):
                            _block_info = f"\n检测到拦截类型: {_recheck.get('type', 'unknown')}"
                        alert_generic(
                            url,
                            f"页面处理超时（{_elapsed:.0f}s > {EXTRACTION_TIMEOUT_SEC}s），"
                            f"可能存在隐形验证码或页面加载异常。{_block_info}",
                            screenshot_path=ss_path,
                        )
                        rec["status"] = "TIMEOUT"
                        data = None
                        break

                    print("  extracting...")
                    data = _run_js(ws_url, _JS)
                    if isinstance(data, dict) and not _shein_page_needs_retry(data):
                        break
                    if attempt < PAGE_LOAD_RETRIES - 1:
                        print(
                            f"  transient/error page (try {attempt + 1}/{PAGE_LOAD_RETRIES}), "
                            "reloading..."
                        )
                        time.sleep(RELOAD_PAUSE_SEC)
                        ws_url = _reload_tab_and_wait(CDP_PORT, tab_id)

                if rec.get("status") == "TIMEOUT":
                    print("  [跳过] 页面超时")
                    records.append(rec)
                    _inter_url_pause(i, len(urls))
                    continue  # tab closed by finally

                if not isinstance(data, dict):
                    raise ValueError("JS returned unexpected type — page may not have loaded")

                shipping = _calc_shipping(data)
                price    = data.get("price") or 0.0
                ebay     = _ebay_listing_price(price, shipping)
                thresh   = data.get("free_threshold")

                ship_note = (
                    f"threshold=${thresh:.2f} — price ${price:.2f} "
                    f"{'≥ threshold → FREE' if price >= thresh else '< threshold → $' + str(DEFAULT_SHIPPING_FEE)}"
                    if thresh is not None else
                    "unconditional FREE" if data.get("unconditional_free") else
                    f"no shipping info → default ${DEFAULT_SHIPPING_FEE}"
                )

                sku_prices = data.get("sku_prices") or []
                rec.update({
                    "sku":            data.get("goods_sn") or data.get("goods_id", ""),
                    "price":          price,
                    "shipping":       shipping,
                    "website":        "us.shein.com",
                    "store_name":     data.get("store_name", ""),
                    "original_title": data.get("title", ""),
                    "title":          data.get("title", ""),
                    "variations":     data.get("variations", {}),
                    "ebay_price":     ebay,
                    "media":          data.get("media", {}),
                    "goods_imgs":     data.get("goods_imgs") or [],
                    "sku_prices":     sku_prices,
                    "main_sale_attrs": data.get("main_sale_attrs") or [],
                    "seq_num":        seq_list[i - 1] if seq_list else start_seq + (i - 1),
                })

                # 如果颜色是单一选中值，追加到标题
                try:
                    vars_ = rec.get("variations") or {}
                    color_key = next((k for k in vars_.keys() if "color" in (k or "").lower()), None)
                    if color_key:
                        cv = vars_.get(color_key) or []
                        if isinstance(cv, list) and len(cv) == 1:
                            color_val = _clean_ws(str(cv[0]))
                            if color_val and color_val.lower() not in (rec["title"] or "").lower():
                                rec["title"] = _clean_ws(f"{rec['title']} - {color_val}")
                except Exception:
                    pass
                if not rec["title"]:
                    rec["status"] = "PARSE_ERROR"

                # 从 sku_prices attrs 补充 variations 中缺失的属性（如 Size: Double）
                if sku_prices:
                    for sp in sku_prices:
                        for ak, av in (sp.get("attrs") or {}).items():
                            if not ak or not av:
                                continue
                            exists = any(vk.lower() == ak.lower() for vk in rec["variations"])
                            if not exists:
                                all_vals = []
                                seen_v = set()
                                for sp2 in sku_prices:
                                    v2 = (sp2.get("attrs") or {}).get(ak, "")
                                    if v2 and v2 not in seen_v:
                                        seen_v.add(v2)
                                        all_vals.append(v2)
                                if all_vals:
                                    rec["variations"][ak] = all_vals

                # 生成 eBay title：AI 优先，失败 fallback 规则
                _orig_t = rec.get("original_title", "")
                _ai_title = _make_ebay_title_ai(_orig_t)
                if _ai_title:
                    rec["ebay_title"] = _ai_title
                    print(f"  eBay title : {_ai_title} (AI)")
                else:
                    rec["ebay_title"] = _make_ebay_title(
                        _orig_t, rec.get("variations", {})
                    )
                    print(f"  eBay title : {rec['ebay_title']} (fallback)")

                seq = rec["seq_num"]
                goods_imgs_count = len(rec.get("goods_imgs") or [])
                print(f"  seq        : {seq}")
                print(f"  title      : {rec['title'][:65]}")
                print(f"  sku        : {rec['sku']}")
                print(f"  price      : ${price:.2f}")
                print(f"  shipping   : ${shipping:.2f}  ({ship_note})")
                print(f"  store      : {rec['store_name']}")
                print(f"  eBay price : ${ebay:.2f}")
                print(f"  variations : {rec['variations']}")
                print(f"  sku_prices : {len(sku_prices)} 个变体")
                print(f"  goods_imgs : {goods_imgs_count} 张（JSON解析）")

                base_dir = Path.cwd()
                media_folder = _download_media(rec, base_dir, seq_num=seq)
                _write_ebay_listing_txt(rec, media_folder)
                first_img = _first_product_image_path(media_folder)
                rec["first_image_path"] = str(first_img) if first_img else ""
                if media_folder:
                    n_webp = len(list(Path(media_folder).glob("img_*.webp")))
                    n_img  = len(list(Path(media_folder).glob("img_*.*")))
                    n_vid  = len(list(Path(media_folder).glob("video_*.mp4")))
                    print(f"  media      : {n_img} image(s) ({n_webp} webp), {n_vid} video(s)")

            except Exception as e:
                rec["status"] = f"ERROR: {e}"
                print(f"  ERROR: {e}")
            finally:
                # Each URL got its own tab — close it to avoid Chrome accumulation
                if tab_id is not None:
                    _close_tab(CDP_PORT, tab_id)

            records.append(rec)

            # 限流检测：连续失败 N 次则停止
            if rec.get("status") != "OK":
                _consecutive_fails += 1
                if _consecutive_fails >= RATE_LIMIT_CONSECUTIVE:
                    print(f"\n  [限流] 连续 {_consecutive_fails} 个 URL 失败，判定为 Shein 限流，停止当前批次")
                    for j in range(i + 1, len(urls) + 1):
                        if j <= len(urls):
                            records.append({
                                "url": urls[j - 1],
                                "status": "RATE_LIMITED",
                                "seq_num": seq_list[j - 1] if seq_list else start_seq + (j - 1),
                            })
                    _rate_limited = True
                    break
            else:
                _consecutive_fails = 0

            _inter_url_pause(i, len(urls))

    finally:
        if launched_new and chrome_proc is not None and not KEEP_CHROME_OPEN:
            chrome_proc.terminate()
            print("\nChrome closed.")
        elif launched_new and chrome_proc is not None and KEEP_CHROME_OPEN:
            print("\nChrome left running (persistent session).")
        else:
            print("\nChrome left running (reused existing session).")

    expanded = _expand_records(records)
    _save_excel(expanded, output)
    print(f"\nSaved to '{output}'  ({len(expanded)} row(s), from {len(records)} product(s))")

    # _retry.txt 已停用 — 失败记录直接在输入 Excel 中标 Failed，下次手动重跑
    # failed = [r for r in records if r.get("status") not in ("OK", None)]
    # retry_path = Path(output).parent / "_retry.txt"
    # if failed:
    #     lines = []
    #     for r in failed:
    #         seq = r.get("seq_num", "?")
    #         status = r.get("status", "UNKNOWN")
    #         url = r.get("url", "")
    #         lines.append(f"{seq}\t{status}\t{url}")
    #     retry_path.write_text("\n".join(lines) + "\n", encoding="utf-8")
    #     print(f"  [Retry] {len(failed)} 个失败 URL 已写入 {retry_path.name}")
    # else:
    #     if retry_path.exists():
    #         retry_path.unlink(missing_ok=True)

    if _rate_limited:
        raise RateLimitError(f"Shein 限流：连续 {RATE_LIMIT_CONSECUTIVE} 个 URL 失败")

    return records
